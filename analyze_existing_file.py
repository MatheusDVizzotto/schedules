# analyze_existing_file.py - Analyze your existing Google Drive file
from google_drive_handler import GoogleDriveHandler
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

print("=" * 80)
print("Analyzing Your Existing Google Drive File")
print("=" * 80)
print()

# Get file ID
file_id = input("Enter your Google Drive File ID (from URL): ").strip()

if not file_id:
    print("No file ID provided. Exiting.")
    exit()

try:
    print("\nConnecting to Google Drive...")
    gdrive = GoogleDriveHandler()

    print("Downloading file...")
    buffer = gdrive.download_file(file_id)

    if not buffer:
        print("✗ Could not download file")
        exit()

    print("✓ File downloaded")
    print(f"  Size: {len(buffer.getvalue())} bytes")

    print("\nAnalyzing Excel structure...")
    buffer.seek(0)
    wb = load_workbook(buffer)

    print(f"\n{'=' * 80}")
    print("WORKBOOK STRUCTURE")
    print('=' * 80)

    print(f"\nSheets: {wb.sheetnames}")
    print(f"Active sheet: {wb.active.title}")

    ws = wb.active

    # Analyze the structure
    print(f"\n{'=' * 80}")
    print("CONTENT ANALYSIS")
    print('=' * 80)

    # Find machines
    print("\n1. MACHINES (looking in Column A):")
    print("-" * 40)
    machines_found = {}
    for row in range(1, 100):
        cell_value = ws[f'A{row}'].value
        if cell_value:
            machines_found[row] = str(cell_value).strip()

    if machines_found:
        for row, machine in sorted(machines_found.items())[:50]:  # Show first 50
            print(f"   Row {row}: {machine}")
    else:
        print("   No data found in Column A")

    # Find workers (looking in row 1 across many columns)
    print("\n2. WORKERS (looking in Row 1 across columns):")
    print("-" * 40)
    workers_found = {}
    for col in range(1, 100):
        col_letter = get_column_letter(col)
        cell_value = ws[f'{col_letter}1'].value
        if cell_value and str(cell_value).strip():
            workers_found[col] = {
                'letter': col_letter,
                'name': str(cell_value).strip()
            }

    if workers_found:
        for col, worker_info in sorted(workers_found.items())[:50]:
            print(f"   Column {worker_info['letter']} ({col}): {worker_info['name']}")
    else:
        print("   No data found in Row 1")

    # Sample proficiency data
    print("\n3. PROFICIENCY DATA (sample - checking intersections):")
    print("-" * 40)
    sample_machines = list(machines_found.keys())[:5]
    sample_workers = list(workers_found.keys())[:5]

    if sample_machines and sample_workers:
        print(f"\n   Checking intersections of first 5 machines × workers:")
        for machine_row in sample_machines:
            for worker_col in sample_workers:
                col_letter = get_column_letter(worker_col)
                cell_value = ws[f'{col_letter}{machine_row}'].value
                if cell_value:
                    print(f"   {col_letter}{machine_row}: {cell_value}")

    # Check for existing schedule sheets
    print("\n4. EXISTING SCHEDULE SHEETS:")
    print("-" * 40)
    schedule_sheets = [s for s in wb.sheetnames if s != wb.active.title]
    if schedule_sheets:
        for sheet in schedule_sheets:
            print(f"   - {sheet}")
    else:
        print("   No additional sheets found")

    # Show overall dimensions
    print(f"\n{'=' * 80}")
    print("DIMENSIONS")
    print('=' * 80)
    print(f"Max Row: {ws.max_row}")
    print(f"Max Column: {ws.max_column} ({get_column_letter(ws.max_column)})")

    # Recommendations
    print(f"\n{'=' * 80}")
    print("RECOMMENDATIONS FOR config.py")
    print('=' * 80)

    if machines_found:
        machine_rows = sorted(machines_found.keys())
        print(f"\nMachine rows found: {machine_rows}")

        # Try to detect ranges
        ranges = []
        current_range_start = None
        prev_row = None

        for row in machine_rows:
            if current_range_start is None:
                current_range_start = row
            elif row != prev_row + 1:
                ranges.append((current_range_start, prev_row))
                current_range_start = row
            prev_row = row

        if current_range_start is not None:
            ranges.append((current_range_start, prev_row))

        print(f"\nDetected machine ranges:")
        print("MACHINE_RANGES = [")
        for start, end in ranges:
            print(f"    ({start}, {end}),  # {machines_found.get(start, '')} to {machines_found.get(end, '')}")
        print("]")

    if workers_found:
        worker_cols = sorted(workers_found.keys())
        first_col = worker_cols[0]
        last_col = worker_cols[-1]
        print(f"\nWorker columns: {first_col} to {last_col}")
        print(f"WORKER_COL_START = {first_col}  # Column {get_column_letter(first_col)}")
        print(f"WORKER_COL_END = {last_col}    # Column {get_column_letter(last_col)}")

    print(f"\n{'=' * 80}")
    print("NEXT STEPS")
    print('=' * 80)
    print("\n1. Update config.py with the above ranges")
    print(f"2. Set: GOOGLE_DRIVE_FILE_ID = '{file_id}'")
    print("3. Set: USE_FILE_ID = True")
    print("\n4. Run: python app.py")

    # Save analysis to file
    with open('file_analysis.txt', 'w', encoding='utf-8') as f:
        f.write(f"File ID: {file_id}\n")
        f.write(f"Sheets: {wb.sheetnames}\n\n")
        f.write("Machines:\n")
        for row, machine in sorted(machines_found.items()):
            f.write(f"  Row {row}: {machine}\n")
        f.write("\nWorkers:\n")
        for col, worker_info in sorted(workers_found.items()):
            f.write(f"  Column {worker_info['letter']}: {worker_info['name']}\n")
        f.write(f"\nRanges detected:\n")
        if ranges:
            f.write("MACHINE_RANGES = [\n")
            for start, end in ranges:
                f.write(f"    ({start}, {end}),\n")
            f.write("]\n")

    print("\n✓ Analysis saved to: file_analysis.txt")

except Exception as e:
    print(f"\n✗ Error: {e}")
    import traceback

    traceback.print_exc()
    print("\nMake sure:")
    print("1. File ID is correct")
    print("2. You have access to the file")
    print("3. File is a valid Excel file (.xlsx)")