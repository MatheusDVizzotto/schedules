# create_template_gdrive.py - Create template Excel file and upload to Google Drive
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
from google_drive_handler import GoogleDriveHandler
import io
import random


def create_template():
    """Create template Excel file with sample data"""
    print("Creating template Excel file...")

    wb = Workbook()
    ws = wb.active
    ws.title = "Main"

    # Add sample machines
    print("Adding sample machines...")
    machines = ['CNC Machine 1', 'CNC Machine 2', 'Lathe 1', 'Lathe 2']
    for i, machine in enumerate(machines, start=5):
        ws[f'A{i}'] = machine

    machines2 = ['Mill 1', 'Mill 2', 'Drill Press 1', 'Drill Press 2',
                 'Grinder 1', 'Grinder 2', 'Press 1', 'Press 2',
                 'Welder 1', 'Welder 2', 'Assembly 1', 'Assembly 2', 'Quality Control']
    for i, machine in enumerate(machines2, start=10):
        ws[f'A{i}'] = machine

    machines3 = ['Packaging 1', 'Packaging 2', 'Shipping 1', 'Shipping 2',
                 'Maintenance 1', 'Maintenance 2', 'Storage 1', 'Storage 2',
                 'Office 1', 'Office 2', 'Break Room', 'Training Room', 'Conference Room']
    for i, machine in enumerate(machines3, start=24):
        ws[f'A{i}'] = machine

    # Add sample workers
    print("Adding sample workers...")
    workers = ['John Smith', 'Jane Doe', 'Bob Johnson', 'Alice Williams',
               'Charlie Brown', 'Diana Prince', 'Eve Martinez', 'Frank Garcia',
               'Grace Lee', 'Henry Wilson']
    for col_num, worker in enumerate(workers, start=41):
        col_letter = get_column_letter(col_num)
        ws[f'{col_letter}1'] = worker

    # Add sample proficiency data
    print("Adding sample proficiency data...")
    proficiency_levels = ['Main Role', 'Competent', 'Trainee']

    random.seed(42)

    for row in range(5, 37):
        if ws[f'A{row}'].value:
            num_workers = random.randint(2, 5)
            assigned_cols = random.sample(range(41, 51), num_workers)

            for col in assigned_cols:
                col_letter = get_column_letter(col)
                if col == assigned_cols[0]:
                    ws[f'{col_letter}{row}'] = 'Main Role'
                else:
                    ws[f'{col_letter}{row}'] = random.choice(['Competent', 'Trainee'])

    # Add formatting
    print("Adding formatting...")

    # Header row for workers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col in range(41, 51):
        col_letter = get_column_letter(col)
        cell = ws[f'{col_letter}1']
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Machine names formatting
    machine_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    machine_font = Font(bold=True)

    for row in range(5, 37):
        if ws[f'A{row}'].value:
            cell = ws[f'A{row}']
            cell.fill = machine_fill
            cell.font = machine_font

    # Set column widths
    ws.column_dimensions['A'].width = 20
    for col in range(41, 51):
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = 15

    return wb


def upload_to_gdrive(workbook, filename='schedule.xlsx'):
    """Upload workbook to Google Drive"""
    try:
        print(f"\nConnecting to Google Drive...")
        gdrive = GoogleDriveHandler()

        # Check if file already exists
        existing_file_id = gdrive.get_file_id_by_name(filename)

        if existing_file_id:
            print(f"\n⚠ File '{filename}' already exists in Google Drive")
            overwrite = input("Do you want to overwrite it? (yes/no): ")
            if overwrite.lower() != 'yes':
                print("Upload cancelled")
                return None

        # Save workbook to buffer
        print("Preparing file for upload...")
        file_buffer = io.BytesIO()
        workbook.save(file_buffer)
        file_buffer.seek(0)

        if existing_file_id:
            # Update existing file
            print("Updating existing file...")
            result = gdrive.upload_file(existing_file_id, file_buffer)
            file_id = existing_file_id
        else:
            # Create new file
            print("Uploading new file...")
            file_id = gdrive.create_file(filename, file_buffer)

        if file_id:
            print(f"\n✓ File uploaded successfully!")
            print(f"  File ID: {file_id}")
            print(f"  Filename: {filename}")

            # Get file metadata
            metadata = gdrive.get_file_metadata(file_id)
            if metadata and 'webViewLink' in metadata:
                print(f"  View in browser: {metadata['webViewLink']}")

            # Update config
            update_config(file_id, filename)

            return file_id
        else:
            print("\n❌ Upload failed")
            return None

    except FileNotFoundError:
        print("\n❌ Error: credentials.json not found")
        print("Please run: python setup_google_drive.py")
        return None
    except Exception as e:
        print(f"\n❌ Error uploading to Google Drive: {e}")
        return None


def update_config(file_id, filename):
    """Update config.py with file ID"""
    try:
        with open('config.py', 'r') as f:
            content = f.read()

        import re
        content = re.sub(
            r"GOOGLE_DRIVE_FILE_ID = ['\"].*?['\"]",
            f"GOOGLE_DRIVE_FILE_ID = '{file_id}'",
            content
        )

        content = re.sub(
            r"USE_FILE_ID = (True|False)",
            "USE_FILE_ID = True",
            content
        )

        content = re.sub(
            r"GOOGLE_DRIVE_FILENAME = ['\"].*?['\"]",
            f"GOOGLE_DRIVE_FILENAME = '{filename}'",
            content
        )

        with open('config.py', 'w') as f:
            f.write(content)

        print(f"\n✓ Updated config.py")

    except Exception as e:
        print(f"\n⚠ Could not auto-update config.py: {e}")
        print("Please manually update config.py with:")
        print(f"  GOOGLE_DRIVE_FILE_ID = '{file_id}'")
        print(f"  USE_FILE_ID = True")


if __name__ == '__main__':
    print("=" * 80)
    print("Google Drive Template Creator")
    print("=" * 80)
    print()

    filename = input("Enter filename for the template (default: schedule.xlsx): ").strip()
    if not filename:
        filename = 'schedule.xlsx'

    if not filename.endswith('.xlsx'):
        filename += '.xlsx'

    print()
    workbook = create_template()

    print("\nTemplate created successfully!")
    print("\nTemplate contains:")
    print("  - 30 sample machines")
    print("  - 10 sample workers")
    print("  - Realistic proficiency assignments")
    print("  - Professional formatting")

    print("\nUploading to Google Drive...")
    file_id = upload_to_gdrive(workbook, filename)

    if file_id:
        print("\n" + "=" * 80)
        print("SUCCESS! Your template is ready.")
        print("=" * 80)
        print("\nNext steps:")
        print("1. Edit the file in Google Drive if needed")
        print("2. Run: python app.py")
        print("3. Open browser: http://localhost:5000")
        print("=" * 80)
    else:
        print("\n" + "=" * 80)
        print("Upload failed. Please check the errors above.")
        print("=" * 80)