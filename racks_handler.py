"""
racks_handler.py — Manages the racks_management Google Drive file.

Each sheet tab in the workbook represents a location.
Bays are saved to the tab matching their selected location.
A hidden '_placeholder_' sheet keeps the workbook valid while no locations exist.
"""
import io
from collections import defaultdict

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from google_drive_handler import GoogleDriveHandler

FILENAME      = 'racks_management'
HEADERS       = ['Bay Code', 'Size Preferable', 'Actual Size', 'Quantity', 'Quantity Unit', 'Item Type', 'Dimensions']
STOCK_SHEET   = '_stock_'
STOCK_HEADERS = ['Size', 'Item Type', 'Dimensions', 'Min On Hand', 'Max On Hand']

HEADER_FILL  = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
THIN_SIDE    = Side(style='thin')
THIN_BORDER  = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
CENTER_ALIGN = Alignment(horizontal='center', vertical='center')


class RacksHandler:

    def __init__(self, schedule_file_id: str | None = None):
        self.gdrive    = GoogleDriveHandler()
        self._folder_id = self._resolve_folder(schedule_file_id)
        self.file_id   = self.gdrive.get_file_id_by_name(FILENAME, self._folder_id)
        self.workbook  = None

    def _resolve_folder(self, schedule_file_id: str | None) -> str | None:
        if not schedule_file_id:
            return None
        meta = self.gdrive.get_file_metadata(schedule_file_id)
        if meta and meta.get('parents'):
            return meta['parents'][0]
        return None

    # ------------------------------------------------------------------

    def load(self):
        if self.file_id:
            buf           = self.gdrive.download_file(self.file_id)
            self.workbook = openpyxl.load_workbook(buf)
        else:
            self.workbook = self._new_workbook()
            buf           = self._serialise()
            self.file_id  = self.gdrive.create_file(FILENAME, buf, folder_id=self._folder_id)
            print(f"  Created new Google Drive file '{FILENAME}' — id={self.file_id}")

    def close(self):
        if self.workbook:
            self.workbook.close()
            self.workbook = None

    # ------------------------------------------------------------------
    # Locations
    # ------------------------------------------------------------------

    def get_locations(self) -> list[str]:
        """Return all location sheet names (excludes internal placeholder sheet)."""
        return [s for s in self.workbook.sheetnames if not s.startswith('_')]

    def add_location(self, name: str) -> list[str]:
        """Create a new location tab with headers and save. Returns updated location list."""
        name = name.strip()
        if not name:
            raise ValueError("Location name cannot be empty")
        if name in self.workbook.sheetnames:
            raise ValueError(f"Location '{name}' already exists")
        self._create_location_sheet(name)
        # Remove the placeholder once at least one real location exists
        if '_placeholder_' in self.workbook.sheetnames:
            del self.workbook['_placeholder_']
        self.gdrive.upload_file(self.file_id, self._serialise())
        return self.get_locations()

    # ------------------------------------------------------------------
    # Racks
    # ------------------------------------------------------------------

    def get_racks(self) -> list[dict]:
        """Return all racks across all location tabs, each with a 'location' field."""
        racks = []
        for loc in self.get_locations():
            ws = self.workbook[loc]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if all(v is None for v in row):
                    continue
                racks.append({
                    'location':        loc,
                    'bay_code':        str(row[0] or '').strip(),
                    'size_preferable': str(row[1] or '').strip(),
                    'actual_size':     str(row[2] or '').strip(),
                    'quantity':        str(row[3] or '').strip(),
                    'qty_unit':        str(row[4] or '').strip() if len(row) > 4 else '',
                    'item_type':       str(row[5] or '').strip() if len(row) > 5 else '',
                    'item_subtype':    str(row[6] or '').strip() if len(row) > 6 else '',
                })
        return racks

    def save_racks(self, racks: list[dict]):
        """Group racks by location and overwrite each location tab."""
        by_location = defaultdict(list)
        for rack in racks:
            loc = rack.get('location', '').strip()
            if loc:
                by_location[loc].append(rack)

        # Clear all location sheets
        for loc in self.get_locations():
            ws = self.workbook[loc]
            if ws.max_row > 1:
                ws.delete_rows(2, ws.max_row - 1)

        # Write rows to their sheets
        for loc, loc_racks in by_location.items():
            if loc not in self.workbook.sheetnames:
                continue
            ws = self.workbook[loc]
            for rack in loc_racks:
                row_idx = ws.max_row + 1
                values  = [
                    rack.get('bay_code', ''),
                    rack.get('size_preferable', ''),
                    rack.get('actual_size', ''),
                    rack.get('quantity', ''),
                    rack.get('qty_unit', ''),
                    rack.get('item_type', ''),
                    rack.get('item_subtype', ''),
                ]
                for col_idx, val in enumerate(values, start=1):
                    cell           = ws.cell(row=row_idx, column=col_idx, value=val or None)
                    cell.border    = THIN_BORDER
                    cell.alignment = CENTER_ALIGN

        self.gdrive.upload_file(self.file_id, self._serialise())

    # ------------------------------------------------------------------
    # Stock
    # ------------------------------------------------------------------

    def get_stock(self, racks: list[dict] | None = None) -> list[dict]:
        """Return all stock items with computed qty_on_hand from bays."""
        ws = self._ensure_stock_sheet()
        if racks is None:
            racks = self.get_racks()
        items = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(v is None for v in row):
                continue
            size       = str(row[0] or '').strip()
            item_type  = str(row[1] or '').strip()
            dimensions = str(row[2] or '').strip()
            if dimensions == 'All Dimensions':
                qty_on_hand = sum(
                    float(r['quantity']) for r in racks
                    if r['actual_size'] == size
                    and r['item_type']  == item_type
                    and r['quantity']
                )
            else:
                qty_on_hand = sum(
                    float(r['quantity']) for r in racks
                    if r['actual_size']   == size
                    and r['item_type']    == item_type
                    and r['item_subtype'] == dimensions
                    and r['quantity']
                )
            items.append({
                'size':        size,
                'item_type':   item_type,
                'dimensions':  dimensions,
                'qty_on_hand': qty_on_hand,
                'min_on_hand': str(row[3] or '').strip(),
                'max_on_hand': str(row[4] or '').strip(),
            })
        return items

    def save_stock(self, items: list[dict]):
        """Overwrite the stock sheet with the provided items."""
        ws = self._ensure_stock_sheet()
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)
        for item in items:
            row_idx = ws.max_row + 1
            values  = [
                item.get('size', ''),
                item.get('item_type', ''),
                item.get('dimensions', ''),
                item.get('min_on_hand', ''),
                item.get('max_on_hand', ''),
            ]
            for col_idx, val in enumerate(values, start=1):
                cell           = ws.cell(row=row_idx, column=col_idx, value=val or None)
                cell.border    = THIN_BORDER
                cell.alignment = CENTER_ALIGN
        self.gdrive.upload_file(self.file_id, self._serialise())

    def _ensure_stock_sheet(self):
        if STOCK_SHEET not in self.workbook.sheetnames:
            ws = self.workbook.create_sheet(STOCK_SHEET)
            for col, title in enumerate(STOCK_HEADERS, start=1):
                cell           = ws.cell(row=1, column=col, value=title)
                cell.fill      = HEADER_FILL
                cell.font      = Font(bold=True)
                cell.alignment = CENTER_ALIGN
                cell.border    = THIN_BORDER
            ws.column_dimensions['A'].width = 16
            ws.column_dimensions['B'].width = 14
            ws.column_dimensions['C'].width = 18
            ws.column_dimensions['D'].width = 14
            ws.column_dimensions['E'].width = 14
        return self.workbook[STOCK_SHEET]

    # ------------------------------------------------------------------

    def _create_location_sheet(self, name: str):
        ws = self.workbook.create_sheet(name)
        for col, title in enumerate(HEADERS, start=1):
            cell           = ws.cell(row=1, column=col, value=title)
            cell.fill      = HEADER_FILL
            cell.font      = Font(bold=True)
            cell.alignment = CENTER_ALIGN
            cell.border    = THIN_BORDER
        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 16
        ws.column_dimensions['F'].width = 14
        ws.column_dimensions['G'].width = 18
        return ws

    def _new_workbook(self) -> openpyxl.Workbook:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '_placeholder_'  # hidden from UI; removed when first location is added
        return wb

    def _serialise(self) -> io.BytesIO:
        buf = io.BytesIO()
        self.workbook.save(buf)
        buf.seek(0)
        return buf
