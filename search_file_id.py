# search_file_id.py - Search for files in Google Drive and get File IDs
from google_drive_handler import GoogleDriveHandler
from datetime import datetime
import re


def format_size(size_bytes):
    """Format file size in human-readable format"""
    if not size_bytes:
        return "Unknown"

    size = int(size_bytes)
    for unit in ['B', 'KB', 'MB', 'GB']:
        if size < 1024.0:
            return f"{size:.2f} {unit}"
        size /= 1024.0
    return f"{size:.2f} TB"


def format_date(date_str):
    """Format ISO date to readable format"""
    try:
        dt = datetime.fromisoformat(date_str.replace('Z', '+00:00'))
        return dt.strftime('%Y-%m-%d %H:%M')
    except:
        return date_str


def search_files(gdrive, query=None, mime_type=None, name_contains=None):
    """Search for files with filters"""

    # Build query
    query_parts = ["trashed=false"]

    if mime_type:
        query_parts.append(f"mimeType='{mime_type}'")

    if name_contains:
        query_parts.append(f"name contains '{name_contains}'")

    if query:
        query_parts.append(query)

    full_query = " and ".join(query_parts)

    print(f"Searching with query: {full_query}")
    print()

    try:
        results = gdrive.service.files().list(
            q=full_query,
            pageSize=100,
            fields="files(id, name, mimeType, size, modifiedTime, webViewLink, owners, shared)",
            orderBy="modifiedTime desc"
        ).execute()

        return results.get('files', [])
    except Exception as e:
        print(f"Error searching: {e}")
        return []


def display_files(files):
    """Display files in a nice format"""
    if not files:
        print("No files found.")
        return None

    print(f"Found {len(files)} file(s):")
    print("=" * 100)

    for i, file in enumerate(files, 1):
        print(f"\n{i}. {file['name']}")
        print(f"   {'─' * 95}")
        print(f"   File ID:       {file['id']}")
        print(f"   Type:          {file.get('mimeType', 'Unknown')}")
        print(f"   Size:          {format_size(file.get('size'))}")
        print(f"   Modified:      {format_date(file.get('modifiedTime', 'Unknown'))}")
        print(f"   Shared:        {'Yes' if file.get('shared') else 'No'}")

        if 'owners' in file and file['owners']:
            owner = file['owners'][0].get('displayName', file['owners'][0].get('emailAddress', 'Unknown'))
            print(f"   Owner:         {owner}")

        if 'webViewLink' in file:
            print(f"   Link:          {file['webViewLink']}")

    print("\n" + "=" * 100)
    return files


def interactive_search():
    """Interactive file search"""
    print("=" * 100)
    print("Google Drive File ID Finder")
    print("=" * 100)
    print()

    try:
        print("Connecting to Google Drive...")
        gdrive = GoogleDriveHandler()
        print("✓ Connected successfully")
        print()

        while True:
            print("\n" + "─" * 100)
            print("SEARCH OPTIONS")
            print("─" * 100)
            print("1. Search all Excel files (.xlsx)")
            print("2. Search all Google Sheets")
            print("3. Search by filename")
            print("4. Search all files (recent 100)")
            print("5. Search with custom query")
            print("6. List files in a specific folder")
            print("0. Exit")
            print()

            choice = input("Choose an option (0-6): ").strip()

            if choice == '0':
                print("\nExiting...")
                break

            files = None

            if choice == '1':
                print("\nSearching for Excel files (.xlsx)...")
                files = search_files(
                    gdrive,
                    mime_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )

            elif choice == '2':
                print("\nSearching for Google Sheets...")
                files = search_files(
                    gdrive,
                    mime_type='application/vnd.google-apps.spreadsheet'
                )

            elif choice == '3':
                filename = input("\nEnter filename (or part of it): ").strip()
                if filename:
                    print(f"\nSearching for files containing '{filename}'...")
                    files = search_files(gdrive, name_contains=filename)

            elif choice == '4':
                print("\nListing recent files...")
                files = search_files(gdrive)

            elif choice == '5':
                print("\nCustom query examples:")
                print("  - name = 'exact_filename.xlsx'")
                print("  - fullText contains 'search text'")
                print("  - modifiedTime > '2024-01-01T00:00:00'")
                print()
                custom_query = input("Enter custom query: ").strip()
                if custom_query:
                    print(f"\nSearching...")
                    files = search_files(gdrive, query=custom_query)

            elif choice == '6':
                folder_id = input("\nEnter folder ID: ").strip()
                if folder_id:
                    print(f"\nListing files in folder...")
                    try:
                        results = gdrive.service.files().list(
                            q=f"'{folder_id}' in parents and trashed=false",
                            pageSize=100,
                            fields="files(id, name, mimeType, size, modifiedTime, webViewLink)"
                        ).execute()
                        files = results.get('files', [])
                    except Exception as e:
                        print(f"Error: {e}")

            else:
                print("\n✗ Invalid option")
                continue

            # Display results
            if files is not None:
                displayed_files = display_files(files)

                if displayed_files:
                    # Option to select a file
                    print()
                    action = input("Enter file number to copy ID, 'a' to analyze file, or Enter to continue: ").strip()

                    if action.lower() == 'a':
                        file_num = input("Which file number to analyze? ").strip()
                        if file_num.isdigit():
                            idx = int(file_num) - 1
                            if 0 <= idx < len(displayed_files):
                                selected_file = displayed_files[idx]
                                analyze_file(gdrive, selected_file)

                    elif action.isdigit():
                        idx = int(action) - 1
                        if 0 <= idx < len(displayed_files):
                            selected_file = displayed_files[idx]
                            print(f"\n{'=' * 100}")
                            print("SELECTED FILE")
                            print('=' * 100)
                            print(f"Name:     {selected_file['name']}")
                            print(f"File ID:  {selected_file['id']}")
                            print('=' * 100)

                            # Update config option
                            update = input("\nUpdate config.py with this File ID? (yes/no): ").strip().lower()
                            if update == 'yes':
                                update_config_with_file_id(selected_file['id'], selected_file['name'])

    except FileNotFoundError:
        print("\n✗ Error: credentials.json not found")
        print("Please run: python setup_google_drive.py")
    except Exception as e:
        print(f"\n✗ Error: {e}")
        import traceback
        traceback.print_exc()


def analyze_file(gdrive, file):
    """Analyze file structure"""
    print(f"\n{'=' * 100}")
    print(f"ANALYZING: {file['name']}")
    print('=' * 100)

    file_id = file['id']
    mime_type = file.get('mimeType', '')

    # Check if it's an Excel file
    if 'spreadsheet' in mime_type.lower() or file['name'].endswith('.xlsx'):
        try:
            from openpyxl import load_workbook
            from openpyxl.utils import get_column_letter

            print("\nDownloading file...")
            buffer = gdrive.download_file(file_id)

            if not buffer:
                print("✗ Could not download file")
                return

            print("Loading workbook...")
            buffer.seek(0)
            wb = load_workbook(buffer)

            print(f"\n✓ File loaded successfully")
            print(f"  Sheets: {', '.join(wb.sheetnames)}")

            ws = wb.active
            print(f"  Active sheet: {ws.title}")
            print(f"  Dimensions: {ws.max_row} rows × {ws.max_column} columns")

            # Quick preview
            print("\nFirst 10 rows of Column A:")
            print("─" * 50)
            for row in range(1, min(11, ws.max_row + 1)):
                value = ws[f'A{row}'].value
                if value:
                    print(f"  A{row}: {value}")

            print("\nFirst 10 columns of Row 1:")
            print("─" * 50)
            for col in range(1, min(11, ws.max_column + 1)):
                col_letter = get_column_letter(col)
                value = ws[f'{col_letter}1'].value
                if value:
                    print(f"  {col_letter}1: {value}")

            # Suggest running detailed analysis
            print(f"\n{'=' * 100}")
            print("For detailed analysis, run:")
            print(f"  python analyze_existing_file.py")
            print(f"  Enter File ID: {file_id}")
            print('=' * 100)

        except Exception as e:
            print(f"\n✗ Error analyzing file: {e}")
    else:
        print(f"\n⚠ File type '{mime_type}' cannot be analyzed")
        print("Only Excel (.xlsx) and Google Sheets are supported")


def update_config_with_file_id(file_id, filename):
    """Update config.py with the selected File ID"""
    try:
        with open('config.py', 'r', encoding='utf-8') as f:
            content = f.read()

        # Update FILE_ID
        content = re.sub(
            r"GOOGLE_DRIVE_FILE_ID = ['\"].*?['\"]",
            f"GOOGLE_DRIVE_FILE_ID = '{file_id}'",
            content
        )

        # Update FILENAME
        content = re.sub(
            r"GOOGLE_DRIVE_FILENAME = ['\"].*?['\"]",
            f"GOOGLE_DRIVE_FILENAME = '{filename}'",
            content
        )

        # Set USE_FILE_ID to True
        content = re.sub(
            r"USE_FILE_ID = (True|False)",
            "USE_FILE_ID = True",
            content
        )

        with open('config.py', 'w', encoding='utf-8') as f:
            f.write(content)

        print(f"\n{'=' * 100}")
        print("✓ config.py updated successfully!")
        print('=' * 100)
        print(f"GOOGLE_DRIVE_FILE_ID = '{file_id}'")
        print(f"GOOGLE_DRIVE_FILENAME = '{filename}'")
        print("USE_FILE_ID = True")
        print('=' * 100)

        # Ask if they want to analyze
        analyze = input("\nRun detailed analysis on this file? (yes/no): ").strip().lower()
        if analyze == 'yes':
            print(f"\nRun: python analyze_existing_file.py")
            print(f"File ID: {file_id}")

    except Exception as e:
        print(f"\n✗ Error updating config.py: {e}")
        print("\nPlease manually update config.py with:")
        print(f"  GOOGLE_DRIVE_FILE_ID = '{file_id}'")
        print(f"  USE_FILE_ID = True")


def quick_search_by_name(name):
    """Quick search by exact or partial name"""
    try:
        gdrive = GoogleDriveHandler()
        files = search_files(gdrive, name_contains=name)
        return display_files(files)
    except Exception as e:
        print(f"Error: {e}")
        return None


if __name__ == '__main__':
    import sys

    # Check for command line argument
    if len(sys.argv) > 1:
        search_term = ' '.join(sys.argv[1:])
        print(f"Quick search for: {search_term}")
        print()
        quick_search_by_name(search_term)
    else:
        interactive_search()