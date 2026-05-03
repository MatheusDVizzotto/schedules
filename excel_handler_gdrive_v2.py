# excel_handler_gdrive_v2.py - Preserves original Excel layout
import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from datetime import datetime, timedelta
import io
from google_drive_handler import GoogleDriveHandler
from config import MACHINE_RANGES, WORKER_COL_START, WORKER_COL_END


class ExcelHandlerGDrive:
    def __init__(self, file_id=None, filename=None):
        self.gdrive = GoogleDriveHandler()
        self.file_id = file_id
        self.filename = filename
        self.workbook = None

        # Define fill styles
        self.yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        self.clear_fill = PatternFill(fill_type=None)

        if not self.file_id and self.filename:
            self.file_id = self.gdrive.get_file_id_by_name(self.filename)
            if not self.file_id:
                raise FileNotFoundError(f"File '{self.filename}' not found in Google Drive")

    def load(self):
        """Load the Excel workbook from Google Drive"""
        if not self.file_id:
            raise ValueError("No file_id specified")

        print("Downloading file from Google Drive...")
        file_buffer = self.gdrive.download_file(self.file_id)

        if file_buffer:
            self.workbook = load_workbook(file_buffer)
            print(f"✓ File loaded: {self.workbook.sheetnames}")
            return self.workbook
        else:
            raise Exception("Failed to download file from Google Drive")

    def save(self):
        """Save the Excel workbook back to Google Drive"""
        if not self.workbook:
            raise Exception("No workbook loaded")

        if not self.file_id:
            raise ValueError("No file_id specified")

        print("Preparing to upload to Google Drive...")
        file_buffer = io.BytesIO()
        self.workbook.save(file_buffer)

        print("Uploading to Google Drive...")
        result = self.gdrive.upload_file(self.file_id, file_buffer)

        if not result:
            raise Exception("Failed to upload file to Google Drive")

        print("✓ File uploaded successfully")
        return result

    def get_machines(self, sheet_name=None):
        """Extract machines using ranges from config"""
        if not self.workbook:
            self.load()

        try:
            sheet = self.workbook[sheet_name] if sheet_name else self.workbook.active
        except KeyError:
            sheet = self.workbook.active

        machines = []

        for start, end in MACHINE_RANGES:
            for row in range(start, end + 1):
                try:
                    cell_value = sheet[f'A{row}'].value
                    if cell_value:
                        machines.append({
                            'row': row,
                            'name': str(cell_value).strip()
                        })
                except Exception as e:
                    print(f"Warning: Could not read cell A{row}: {e}")
                    continue

        return machines

    def get_workers(self, sheet_name=None):
        """Extract workers using column range from config"""
        if not self.workbook:
            self.load()

        try:
            sheet = self.workbook[sheet_name] if sheet_name else self.workbook.active
        except KeyError:
            sheet = self.workbook.active

        workers = []

        for col in range(WORKER_COL_START, WORKER_COL_END + 1):
            try:
                col_letter = get_column_letter(col)
                cell_value = sheet[f'{col_letter}1'].value
                if cell_value:
                    workers.append({
                        'col': col,
                        'col_letter': col_letter,
                        'name': str(cell_value).strip()
                    })
            except Exception as e:
                print(f"Warning: Could not read column {col}: {e}")
                continue

        return workers

    def get_proficiency(self, machine_row, worker_col, sheet_name=None):
        """Get proficiency level for a machine-worker combination"""
        if not self.workbook:
            self.load()

        try:
            sheet = self.workbook[sheet_name] if sheet_name else self.workbook.active
        except KeyError:
            sheet = self.workbook.active

        try:
            col_letter = get_column_letter(worker_col)
            cell_value = sheet[f'{col_letter}{machine_row}'].value
            return cell_value if cell_value else ''
        except Exception as e:
            return ''

    def get_all_proficiencies(self, sheet_name=None):
        """Get all proficiency data as a matrix"""
        if not self.workbook:
            self.load()

        try:
            sheet = self.workbook[sheet_name] if sheet_name else self.workbook.active
        except KeyError:
            sheet = self.workbook.active

        machines = self.get_machines(sheet_name)
        workers = self.get_workers(sheet_name)

        proficiencies = {}
        for machine in machines:
            proficiencies[machine['row']] = {}
            for worker in workers:
                try:
                    col_letter = get_column_letter(worker['col'])
                    cell_value = sheet[f'{col_letter}{machine["row"]}'].value
                    proficiencies[machine['row']][worker['col']] = cell_value if cell_value else ''
                except Exception as e:
                    proficiencies[machine['row']][worker['col']] = ''

        return proficiencies

    def update_worker_name(self, worker_col, new_name, sheet_name=None):
        """Update worker name in the header row"""
        if not self.workbook:
            self.load()

        try:
            sheet = self.workbook[sheet_name] if sheet_name else self.workbook.active
        except KeyError:
            sheet = self.workbook.active

        col_letter = get_column_letter(worker_col)
        sheet[f'{col_letter}1'] = new_name
        self.save()

    def update_proficiency(self, machine_row, worker_col, proficiency, sheet_name=None):
        """Update proficiency level for a machine-worker combination"""
        if not self.workbook:
            self.load()

        try:
            sheet = self.workbook[sheet_name] if sheet_name else self.workbook.active
        except KeyError:
            sheet = self.workbook.active

        col_letter = get_column_letter(worker_col)
        sheet[f'{col_letter}{machine_row}'] = proficiency
        self.save()

    def update_proficiencies_bulk(self, proficiencies, sheet_name=None):
        """Update multiple proficiencies at once"""
        if not self.workbook:
            self.load()

        try:
            sheet = self.workbook[sheet_name] if sheet_name else self.workbook.active
        except KeyError:
            sheet = self.workbook.active

        for machine_row, worker_data in proficiencies.items():
            for worker_col, proficiency in worker_data.items():
                col_letter = get_column_letter(worker_col)
                sheet[f'{col_letter}{machine_row}'] = proficiency

        self.save()

    def clear_schedule_data(self, date_str, sheet_name=None):
        """Clear previous schedule data (remove yellow highlights)"""
        if not self.workbook:
            self.load()

        try:
            sheet = self.workbook[sheet_name] if sheet_name else self.workbook.active
        except KeyError:
            sheet = self.workbook.active

        machines = self.get_machines()
        workers = self.get_workers()

        print(f"Clearing schedule for {date_str}...")

        # Clear yellow fills from machine-worker intersections
        for machine in machines:
            for worker in workers:
                col_letter = get_column_letter(worker['col'])
                cell = sheet[f'{col_letter}{machine["row"]}']

                # Only clear if it's a schedule (has yellow fill)
                if cell.fill and cell.fill.start_color.rgb == 'FFFFFF00':
                    cell.value = None
                    cell.fill = self.clear_fill

    def save_schedule_visual(self, date, schedule_data):
        """
        Save schedule preserving original Excel layout
        - Uses existing sheet structure
        - Highlights assigned cells in yellow
        - Adds schedule info to cells
        """
        if not self.workbook:
            self.load()

        # Use active sheet (preserves original structure)
        sheet = self.workbook.active

        print(f"\n{'=' * 80}")
        print(f"Saving schedule for {date.strftime('%Y-%m-%d')}")
        print(f"Sheet: {sheet.title}")
        print('=' * 80)

        # Get structure
        machines = self.get_machines()
        workers = self.get_workers()

        print(f"\nStructure:")
        print(f"  Machines: {len(machines)}")
        print(f"  Workers: {len(workers)}")

        # Clear previous schedule (optional - comment out if you want to keep history)
        # self.clear_schedule_data(date.strftime('%Y-%m-%d'))

        # Process schedule data
        print(f"\nProcessing schedule assignments:")
        print("─" * 80)

        assigned_count = 0

        for entry in schedule_data:
            machine_name = entry.get('machine')
            worker_name = entry.get('worker')
            role = entry.get('role')
            time_start = entry.get('time_start')
            time_finish = entry.get('time_finish')

            # Find machine
            machine = next((m for m in machines if m['name'] == machine_name), None)
            if not machine:
                print(f"  ⚠ Machine not found: {machine_name}")
                continue

            # Find worker
            worker = next((w for w in workers if w['name'] == worker_name), None)
            if not worker:
                print(f"  ⚠ Worker not found: {worker_name}")
                continue

            # Get cell
            col_letter = worker['col_letter']
            row = machine['row']
            cell = sheet[f'{col_letter}{row}']

            # Build schedule text
            schedule_text = f"{time_start}-{time_finish}"
            if role:
                role_abbrev = {
                    'Main Role': 'MR',
                    'Competent': 'C',
                    'Trainee': 'T'
                }.get(role, '')
                if role_abbrev:
                    schedule_text += f" ({role_abbrev})"

            # Update cell
            cell.value = schedule_text
            cell.fill = self.yellow_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.font = Font(size=9)

            print(f"  ✓ {machine_name} → {worker_name}: {schedule_text} at {col_letter}{row}")
            assigned_count += 1

        print("─" * 80)
        print(f"Total assignments: {assigned_count}")
        print()

        # Save to Google Drive
        print("Saving to Google Drive...")
        self.save()

        print("=" * 80)
        print("✓ Schedule saved successfully!")
        print("=" * 80)

    def get_or_create_day_sheet(self, date):
        """Get or create a sheet for specific date (if using multiple sheets)"""
        if not self.workbook:
            self.load()

        sheet_name = date.strftime('%Y-%m-%d')

        if sheet_name in self.workbook.sheetnames:
            return self.workbook[sheet_name]
        else:
            # Create new sheet by copying the main sheet
            source_sheet = self.workbook.active
            new_sheet = self.workbook.copy_worksheet(source_sheet)
            new_sheet.title = sheet_name
            return new_sheet

    def get_schedule(self, date):
        """Get schedule data for a specific date"""
        # This would read back from the cells
        # For now, return empty (implement if needed)
        return []

    def close(self):
        """Close the workbook"""
        if self.workbook:
            self.workbook.close()