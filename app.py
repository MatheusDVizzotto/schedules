"""
app.py - Machine Schedule Manager (Google Drive Edition)
Main Flask application with API endpoints.
"""
# Load .env FIRST — before any imports that read os.environ
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

from flask import Flask, render_template, request, jsonify
from datetime import datetime, timedelta
import json
import os
import threading
import time as _time

from config import (
    GOOGLE_DRIVE_FILE_ID, GOOGLE_DRIVE_FILENAME, USE_FILE_ID,
    MATRIX_FILE_ID, MATRIX_FILENAME, SCHEDULE_FOLDER_ID,
    SECRET_KEY, DEBUG, HOST, PORT, SCHEDULE_MODE
)
from auth import auth_bp, login_required

app = Flask(__name__)
app.config['SECRET_KEY'] = SECRET_KEY
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(days=7)
app.register_blueprint(auth_bp)

# In-memory set of soft-deleted workers; persisted to deleted_workers.json
deleted_workers: set = set()
DELETED_WORKERS_FILE = 'deleted_workers.json'



# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def load_deleted_workers():
    """Load soft-deleted workers from disk."""
    global deleted_workers
    if os.path.exists(DELETED_WORKERS_FILE):
        try:
            with open(DELETED_WORKERS_FILE, 'r') as f:
                deleted_workers = set(json.load(f))
        except (json.JSONDecodeError, IOError) as e:
            print(f"Warning: could not load {DELETED_WORKERS_FILE}: {e}")
            deleted_workers = set()


def save_deleted_workers():
    """Persist soft-deleted workers to disk."""
    try:
        with open(DELETED_WORKERS_FILE, 'w') as f:
            json.dump(list(deleted_workers), f)
    except IOError as e:
        print(f"Warning: could not save {DELETED_WORKERS_FILE}: {e}")




def get_excel_handler():
    """Return a fresh ExcelHandlerGDrive pointing to the Matrix file (not loaded)."""
    from excel_handler_gdrive import ExcelHandlerGDrive
    if MATRIX_FILE_ID:
        return ExcelHandlerGDrive(file_id=MATRIX_FILE_ID)
    return ExcelHandlerGDrive(filename=MATRIX_FILENAME)


# ---------------------------------------------------------------------------
# Workbook cache — avoids re-downloading from Google Drive on every request.
# Read endpoints use this; write endpoints use get_excel_handler() + load()
# directly and call invalidate_cache() after saving.
# ---------------------------------------------------------------------------

_cache_lock = threading.Lock()
_cached_handler = None
_cached_at = 0.0
CACHE_TTL = 10  # seconds


def get_cached_handler():
    """Return a loaded ExcelHandlerGDrive, re-downloading only when stale."""
    global _cached_handler, _cached_at
    with _cache_lock:
        now = _time.monotonic()
        if _cached_handler is None or (now - _cached_at) > CACHE_TTL:
            h = get_excel_handler()
            h.load()
            if _cached_handler is not None:
                _cached_handler.workbook = None  # release old workbook memory
            _cached_handler = h
            _cached_at = now
        return _cached_handler


def invalidate_cache():
    """Force the next read to re-download from Google Drive."""
    global _cached_handler, _cached_at
    with _cache_lock:
        if _cached_handler is not None:
            _cached_handler.workbook = None
        _cached_handler = None
        _cached_at = 0.0


# ---------------------------------------------------------------------------
# Racks workbook cache — same pattern as the schedule cache above.
# ---------------------------------------------------------------------------

_racks_lock = threading.Lock()
_cached_racks_handler = None
_cached_racks_at = 0.0
RACKS_CACHE_TTL = 30  # seconds — shorter to reduce simultaneous memory with master workbook


def get_cached_racks_handler():
    """Return a loaded RacksHandler, re-downloading only when stale."""
    global _cached_racks_handler, _cached_racks_at
    with _racks_lock:
        now = _time.monotonic()
        if _cached_racks_handler is None or (now - _cached_racks_at) > RACKS_CACHE_TTL:
            from racks_handler import RacksHandler
            h = RacksHandler(schedule_file_id=GOOGLE_DRIVE_FILE_ID)
            h.load()
            if _cached_racks_handler is not None:
                _cached_racks_handler.workbook = None
            _cached_racks_handler = h
            _cached_racks_at = now
        return _cached_racks_handler


def get_rack_handler():
    """Return a fresh RacksHandler (not loaded). Use for write operations."""
    from racks_handler import RacksHandler
    return RacksHandler(schedule_file_id=GOOGLE_DRIVE_FILE_ID)


def invalidate_racks_cache():
    """Force the next rack read to re-download from Google Drive."""
    global _cached_racks_handler, _cached_racks_at
    with _racks_lock:
        if _cached_racks_handler is not None:
            _cached_racks_handler.workbook = None
        _cached_racks_handler = None
        _cached_racks_at = 0.0


# ---------------------------------------------------------------------------
# Monthly schedule file management
# Each month gets its own Google Drive file: "schedule MM/YY"
# (e.g. "schedule 05/26" for May 2026).
# ---------------------------------------------------------------------------

_monthly_file_ids: dict = {}   # "MM/YY" → file_id
_monthly_file_lock = threading.Lock()
_schedule_folder_id: str | None = None
_schedule_folder_resolved = False


def _get_schedule_folder() -> str | None:
    """Return the Drive folder ID for monthly schedule files (resolved once)."""
    global _schedule_folder_id, _schedule_folder_resolved
    if _schedule_folder_resolved:
        return _schedule_folder_id
    _schedule_folder_resolved = True
    if SCHEDULE_FOLDER_ID:
        _schedule_folder_id = SCHEDULE_FOLDER_ID
        return _schedule_folder_id
    # Fall back: same folder as the Matrix file
    try:
        from google_drive_handler import GoogleDriveHandler
        gdrive = GoogleDriveHandler()
        fid    = MATRIX_FILE_ID or gdrive.get_file_id_by_name(MATRIX_FILENAME)
        if fid:
            meta    = gdrive.get_file_metadata(fid)
            parents = (meta or {}).get('parents', [])
            _schedule_folder_id = parents[0] if parents else None
    except Exception as e:
        print(f"Warning: could not resolve schedule folder: {e}")
        _schedule_folder_id = None
    return _schedule_folder_id


def _find_monthly_file_id(date) -> str | None:
    """Return the Drive file_id for 'schedule MM/YY', or None if it doesn't exist yet."""
    month_key = date.strftime('%m/%y')
    filename  = f'schedule {month_key}'
    with _monthly_file_lock:
        if month_key in _monthly_file_ids:
            return _monthly_file_ids[month_key]
        from google_drive_handler import GoogleDriveHandler
        gdrive    = GoogleDriveHandler()
        folder_id = _get_schedule_folder()
        file_id   = gdrive.get_file_id_by_name(filename, folder_id=folder_id)
        if file_id:
            _monthly_file_ids[month_key] = file_id
        return file_id


def _register_monthly_file_id(date, file_id: str):
    """Cache a newly created monthly schedule file_id."""
    month_key = date.strftime('%m/%y')
    with _monthly_file_lock:
        _monthly_file_ids[month_key] = file_id


# Per-month schedule cache — mirrors the master-file cache pattern.
_sched_cache: dict = {}   # "MM/YY" → (handler, timestamp)
_sched_cache_lock = threading.Lock()
SCHED_CACHE_TTL = 120  # seconds


def get_cached_schedule_handler(date):
    """Return a loaded handler for the monthly schedule file, or None if no file exists yet."""
    from excel_handler_gdrive import ExcelHandlerGDrive
    month_key = date.strftime('%m/%y')
    with _sched_cache_lock:
        now   = _time.monotonic()
        entry = _sched_cache.get(month_key)
        if entry is None or (now - entry[1]) > SCHED_CACHE_TTL:
            file_id = _find_monthly_file_id(date)
            if not file_id:
                return None   # no schedule saved for this month yet
            h = ExcelHandlerGDrive(file_id=file_id)
            h.load()
            if entry is not None:
                entry[0].workbook = None
            _sched_cache[month_key] = (h, now)
        return _sched_cache.get(month_key, (None,))[0]


def invalidate_schedule_cache(date):
    """Force the next schedule read for this month to re-download."""
    month_key = date.strftime('%m/%y')
    with _sched_cache_lock:
        entry = _sched_cache.pop(month_key, None)
        if entry:
            entry[0].workbook = None


# ---------------------------------------------------------------------------
# Page routes
# ---------------------------------------------------------------------------

@app.route('/')
@login_required
def index():
    return render_template('index.html')

@app.route('/dashboard')
def dashboard():
    return render_template('dashboard.html')

@app.route('/admin')
@login_required
def admin():
    return render_template('admin.html')

@app.route('/workers')
@login_required
def workers_page():
    return render_template('workers.html')

@app.route('/racks')
@login_required
def racks_page():
    return render_template('racks.html')


@app.route('/racks/stock')
@login_required
def stock_page():
    return render_template('stock.html')


@app.route('/racks/stock-dashboard')
@login_required
def stock_dashboard_page():
    return render_template('stock_dashboard.html')


# ---------------------------------------------------------------------------
# API – combined loader (what the frontend actually calls on page load)
# ---------------------------------------------------------------------------

@app.route('/api/workers/all', methods=['GET'])
def get_all_data():
    """
    Single endpoint that returns machines, workers, and proficiencies together.
    The frontend calls this once on page load instead of making three requests.
    """
    try:
        handler = get_cached_handler()
        machines = handler.get_machines()
        all_workers = handler.get_workers()
        proficiencies = handler.get_all_proficiencies()

        active_workers = [w for w in all_workers if w['name'] not in deleted_workers]

        # JSON keys must be strings; convert int row/col keys to strings
        prof_serialisable = {
            str(machine_row): {str(worker_col): val for worker_col, val in worker_data.items()}
            for machine_row, worker_data in proficiencies.items()
        }

        return jsonify({
            'success': True,
            'machines': machines,
            'workers': active_workers,
            'proficiencies': prof_serialisable,
        })
    except Exception as e:
        print(f"Error in get_all_data: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


# ---------------------------------------------------------------------------
# API – machines
# ---------------------------------------------------------------------------

@app.route('/api/machines', methods=['GET'])
def get_machines():
    try:
        handler = get_cached_handler()
        machines = handler.get_machines()
        return jsonify({'success': True, 'machines': machines})
    except Exception as e:
        print(f"Error in get_machines: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


# ---------------------------------------------------------------------------
# API – workers
# ---------------------------------------------------------------------------

@app.route('/api/workers', methods=['GET'])
def get_workers():
    """Return all workers, excluding soft-deleted ones."""
    try:
        handler = get_cached_handler()
        all_workers = handler.get_workers()
        active = [w for w in all_workers if w['name'] not in deleted_workers]
        return jsonify({'success': True, 'workers': active})
    except Exception as e:
        print(f"Error in get_workers: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/workers/delete', methods=['POST'])
def delete_worker():
    """Soft-delete a worker."""
    try:
        data = request.get_json()
        if not data or not data.get('worker_name'):
            return jsonify({'success': False, 'error': 'worker_name is required'}), 400
        name = data['worker_name'].strip()
        deleted_workers.add(name)
        save_deleted_workers()
        return jsonify({'success': True, 'message': f'Worker "{name}" removed from future schedules'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/workers/restore', methods=['POST'])
def restore_worker():
    """Restore a soft-deleted worker."""
    try:
        data = request.get_json()
        if not data or not data.get('worker_name'):
            return jsonify({'success': False, 'error': 'worker_name is required'}), 400
        name = data['worker_name'].strip()
        deleted_workers.discard(name)   # FIX: .remove() raises KeyError if not present
        save_deleted_workers()
        return jsonify({'success': True, 'message': f'Worker "{name}" restored'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/workers/deleted', methods=['GET'])
def get_deleted_workers():
    return jsonify({'success': True, 'workers': list(deleted_workers)})


@app.route('/api/workers/absences', methods=['GET'])
def get_worker_absences():
    try:
        absences = get_cached_handler().get_absences()
        return jsonify({'success': True, 'absences': absences})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/workers/absence/add', methods=['POST'])
def add_worker_absence():
    data      = request.get_json() or {}
    name      = str(data.get('worker_name', '')).strip()
    date_from = str(data.get('date_from', '')).strip()
    date_to   = str(data.get('date_to', '')).strip()
    reason    = str(data.get('reason', '')).strip()
    if not name or not date_from or not date_to:
        return jsonify({'success': False, 'error': 'worker_name, date_from and date_to are required'}), 400
    if date_from > date_to:
        return jsonify({'success': False, 'error': 'date_from must be on or before date_to'}), 400
    try:
        handler  = get_excel_handler()
        handler.load()
        absences = handler.get_absences()
        absences.setdefault(name, []).append({'date_from': date_from, 'date_to': date_to, 'reason': reason})
        handler.save_absences(absences)
        handler.close()
        invalidate_cache()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/workers/absence/remove', methods=['POST'])
def remove_worker_absence():
    data = request.get_json() or {}
    name = str(data.get('worker_name', '')).strip()
    idx  = data.get('index')
    if not name or idx is None:
        return jsonify({'success': False, 'error': 'worker_name and index are required'}), 400
    try:
        handler  = get_excel_handler()
        handler.load()
        absences = handler.get_absences()
        if name not in absences:
            handler.close()
            return jsonify({'success': False, 'error': 'Worker not found'}), 404
        absences[name].pop(int(idx))
        if not absences[name]:
            del absences[name]
        handler.save_absences(absences)
        handler.close()
        invalidate_cache()
        return jsonify({'success': True})
    except (IndexError, ValueError):
        return jsonify({'success': False, 'error': 'Invalid index'}), 400
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/workers/permanent-delete', methods=['POST'])
def permanent_delete_worker():
    """Permanently remove a worker by clearing their column in the spreadsheet."""
    try:
        data = request.get_json()
        if not data or not data.get('worker_name'):
            return jsonify({'success': False, 'error': 'worker_name is required'}), 400
        name = data['worker_name'].strip()
        handler = get_excel_handler()
        handler.load()
        all_workers = handler.get_workers()
        worker = next((w for w in all_workers if w['name'] == name), None)
        if not worker:
            handler.close()
            return jsonify({'success': False, 'error': 'Worker not found'}), 404
        handler.delete_worker_permanently(worker['col'])
        handler.close()
        deleted_workers.discard(name)
        save_deleted_workers()
        invalidate_cache()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/workers/update', methods=['POST'])
def update_worker():
    """Rename a worker in the spreadsheet header row."""
    try:
        data = request.get_json()
        if not data or not data.get('worker_col') or not data.get('new_name'):
            return jsonify({'success': False, 'error': 'worker_col and new_name are required'}), 400
        handler = get_excel_handler()
        handler.load()
        handler.update_worker_name(int(data['worker_col']), data['new_name'].strip())
        handler.close()
        invalidate_cache()
        return jsonify({'success': True, 'message': 'Worker name updated'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/workers/add', methods=['POST'])
def add_worker():
    """
    Add a new worker to the spreadsheet.

    Writes the name into the next empty column in row 1 after the current
    WORKER_COL_END, then increments WORKER_COL_END in config.py so the
    new worker is picked up on the next load.

    Body JSON: { "name": "Worker Name" }
    """
    try:
        data = request.get_json()
        if not data or not data.get('name'):
            return jsonify({'success': False, 'error': 'name is required'}), 400

        name = data['name'].strip()
        if not name:
            return jsonify({'success': False, 'error': 'name cannot be empty'}), 400

        handler = get_excel_handler()
        handler.load()
        new_col = handler.add_worker(name)
        handler.close()
        invalidate_cache()

        # Bump WORKER_COL_END in config.py so the new worker is included next load
        _update_config_worker_col_end(new_col)

        return jsonify({
            'success': True,
            'message': f'Worker "{name}" added at column {new_col}',
            'col': new_col,
            'name': name,
        })
    except Exception as e:
        import traceback
        print(f"Error in add_worker:\n{traceback.format_exc()}")
        return jsonify({'success': False, 'error': str(e)}), 500


def _update_config_worker_col_end(new_col: int):
    """Update WORKER_COL_END in config.py if new_col exceeds current value."""
    import re
    try:
        cfg = open('config.py').read()
        current = int(re.search(r'WORKER_COL_END\s*=\s*(\d+)', cfg).group(1))
        if new_col > current:
            cfg = re.sub(r'(WORKER_COL_END\s*=\s*)\d+', lambda m: m.group(1) + str(new_col), cfg)
            open('config.py', 'w').write(cfg)
            print(f"  Updated WORKER_COL_END: {current} → {new_col}")
    except Exception as e:
        print(f"  Warning: could not update config.py: {e}")


# ---------------------------------------------------------------------------
# API – proficiency
# ---------------------------------------------------------------------------

@app.route('/api/proficiency', methods=['GET'])
def get_proficiency():
    try:
        handler = get_cached_handler()
        data = handler.get_all_proficiencies()
        return jsonify({'success': True, 'proficiency': data})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/proficiency/update', methods=['POST'])
def update_proficiency():
    """Bulk-update proficiency values."""
    try:
        data = request.get_json()
        if not data or 'proficiencies' not in data:
            return jsonify({'success': False, 'error': 'proficiencies payload is required'}), 400
        handler = get_excel_handler()
        handler.load()
        handler.update_proficiencies_bulk(data['proficiencies'])
        handler.close()
        invalidate_cache()
        return jsonify({'success': True, 'message': 'Proficiencies updated'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ---------------------------------------------------------------------------
# API – schedule
# ---------------------------------------------------------------------------

@app.route('/api/schedule/save', methods=['POST'])
def save_schedule():
    """
    Save a day's schedule to Google Drive.

    Body JSON:
        {
            "date": "YYYY-MM-DD",
            "schedule": [
                {
                    "machine":     "Machine Name",
                    "worker":      "Worker Name",
                    "role":        "Main Role" | "Competent" | "Trainee",
                    "time_start":  "HH:MM",
                    "time_finish": "HH:MM"
                }, ...
            ]
        }
    """
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'error': 'No JSON body provided'}), 400

        date_str = data.get('date')
        schedule_data = data.get('schedule', [])

        if not date_str:
            return jsonify({'success': False, 'error': 'date is required (YYYY-MM-DD)'}), 400

        try:
            schedule_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            return jsonify({'success': False, 'error': 'date must be YYYY-MM-DD'}), 400

        if not schedule_data:
            return jsonify({'success': False, 'error': 'schedule list is empty'}), 400

        # Read machine layout from the Matrix file for visual row formatting
        master_layout = get_cached_handler().get_master_layout()

        from excel_handler_gdrive import ExcelHandlerGDrive
        import openpyxl as _xl

        file_id = _find_monthly_file_id(schedule_date)
        handler = ExcelHandlerGDrive(file_id=file_id)

        if file_id:
            # Existing monthly file — download and update
            handler.load()
        else:
            # New month — build workbook in memory; Drive file created on save()
            handler.workbook            = _xl.Workbook()
            handler._create_filename    = f'schedule {schedule_date.strftime("%m/%y")}'
            handler._create_folder_id   = _get_schedule_folder()

        handler.save_schedule_visual(schedule_date, schedule_data, master_layout)
        # save_schedule_visual calls save() internally; for new files self.file_id is set there
        if not file_id:
            _register_monthly_file_id(schedule_date, handler.file_id)

        # Get the real Google Sheets gid via the Sheets API (Google assigns its own
        # large random gids that don't match the xlsx sheetId from openpyxl)
        sheet_gid = None
        if SCHEDULE_MODE == 'NEW_SHEET_PER_DAY':
            sheet_name = schedule_date.strftime('%d-%m-%y')
            sheet_gid = handler.gdrive.get_sheet_gid_api(handler.file_id, sheet_name)

        handler.close()
        invalidate_schedule_cache(schedule_date)

        gid_fragment = f'#gid={sheet_gid}' if sheet_gid is not None else ''
        sheets_url = f'https://docs.google.com/spreadsheets/d/{handler.file_id}/edit{gid_fragment}'
        print(f"  sheet_gid={sheet_gid}  url={sheets_url}")
        return jsonify({
            'success':    True,
            'message':    'Schedule saved successfully to Google Drive',
            'sheets_url': sheets_url,
            'debug_gid':  sheet_gid,
        })
    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        print(f"Error in save_schedule:\n{tb}")
        return jsonify({'success': False, 'error': str(e), 'traceback': tb}), 500


@app.route('/api/schedule/check_date', methods=['GET'])
def check_schedule_date():
    """Return whether a schedule sheet already exists for the given date."""
    date_str = request.args.get('date')
    if not date_str:
        return jsonify({'success': False, 'error': 'date is required'}), 400
    try:
        schedule_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'success': False, 'error': 'date must be YYYY-MM-DD'}), 400

    if SCHEDULE_MODE != 'NEW_SHEET_PER_DAY':
        return jsonify({'success': True, 'exists': False})

    file_id = _find_monthly_file_id(schedule_date)
    if not file_id:
        return jsonify({'success': True, 'exists': False})

    from excel_handler_gdrive import ExcelHandlerGDrive
    handler = ExcelHandlerGDrive(file_id=file_id)
    handler.load()
    sheet_name = schedule_date.strftime('%d-%m-%y')
    exists = sheet_name in handler.workbook.sheetnames
    handler.close()
    return jsonify({'success': True, 'exists': exists})


@app.route('/api/schedule/<date_str>', methods=['GET'])
def get_schedule(date_str):
    try:
        datetime.strptime(date_str, '%Y-%m-%d')
    except ValueError:
        return jsonify({'success': False, 'error': 'date must be YYYY-MM-DD'}), 400
    try:
        date    = datetime.strptime(date_str, '%Y-%m-%d').date()
        handler = get_cached_schedule_handler(date)
        if handler is None:
            return jsonify({'success': True, 'schedule': []})
        schedule = handler.get_schedule(date)
        return jsonify({'success': True, 'schedule': schedule})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ---------------------------------------------------------------------------
# API – racks management
# ---------------------------------------------------------------------------

@app.route('/api/racks', methods=['GET'])
@login_required
def get_racks():
    try:
        handler   = get_cached_racks_handler()
        racks     = handler.get_racks()
        locations = handler.get_locations()
        return jsonify({'success': True, 'racks': racks, 'locations': locations})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/racks/save', methods=['POST'])
@login_required
def save_racks():
    try:
        data = request.get_json()
        if data is None or 'racks' not in data:
            return jsonify({'success': False, 'error': 'racks payload is required'}), 400
        handler = get_rack_handler()
        handler.load()
        handler.save_racks(data['racks'])
        handler.close()
        invalidate_racks_cache()
        return jsonify({'success': True, 'message': 'Racks saved successfully'})
    except Exception as e:
        import traceback
        print(f"Error in save_racks:\n{traceback.format_exc()}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/racks/stock/dashboard', methods=['GET'])
@login_required
def get_stock_dashboard():
    try:
        handler = get_cached_racks_handler()
        bays    = handler.get_racks()
        stock   = handler.get_stock(racks=bays)
        return jsonify({'success': True, 'stock': stock, 'bays': bays})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/racks/stock', methods=['GET'])
@login_required
def get_stock():
    try:
        handler = get_cached_racks_handler()
        items   = handler.get_stock()
        return jsonify({'success': True, 'items': items})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/racks/stock/save', methods=['POST'])
@login_required
def save_stock():
    try:
        data = request.get_json()
        if data is None or 'items' not in data:
            return jsonify({'success': False, 'error': 'items payload is required'}), 400
        handler = get_rack_handler()
        handler.load()
        handler.save_stock(data['items'])
        handler.close()
        invalidate_racks_cache()
        return jsonify({'success': True})
    except Exception as e:
        import traceback
        print(f"Error in save_stock:\n{traceback.format_exc()}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/racks/locations/add', methods=['POST'])
@login_required
def add_rack_location():
    try:
        data = request.get_json()
        if not data or not data.get('name'):
            return jsonify({'success': False, 'error': 'name is required'}), 400
        handler   = get_rack_handler()
        handler.load()
        locations = handler.add_location(data['name'].strip())
        handler.close()
        invalidate_racks_cache()
        return jsonify({'success': True, 'locations': locations})
    except ValueError as e:
        return jsonify({'success': False, 'error': str(e)}), 400
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ---------------------------------------------------------------------------
# API – debug (remove before going to production)
# ---------------------------------------------------------------------------

@app.route('/api/debug/save-test', methods=['POST'])
def debug_save_test():
    """
    Test endpoint — attempts a minimal write to Google Drive and returns
    detailed info about every step so you can pinpoint where it fails.
    """
    import traceback
    result = {
        'step': None,
        'file_id': None,
        'sheet_names_before': [],
        'sheet_names_after': [],
        'sheets_added': [],
        'bytes_written': None,
        'upload_response': None,
        'error': None,
    }
    try:
        result['step'] = '1_get_handler'
        handler = get_excel_handler()
        result['file_id'] = handler.file_id

        result['step'] = '2_load'
        handler.load()
        result['sheet_names_before'] = handler.workbook.sheetnames

        result['step'] = '3_create_sheet'
        import datetime
        test_name = f"DEBUG-{datetime.datetime.now().strftime('%H%M%S')}"
        ws = handler.workbook.create_sheet(test_name)
        ws['A1'] = 'Debug test — safe to delete'
        result['sheet_names_after'] = handler.workbook.sheetnames
        result['sheets_added'] = [s for s in result['sheet_names_after']
                                   if s not in result['sheet_names_before']]

        result['step'] = '4_serialise'
        import io
        buf = io.BytesIO()
        handler.workbook.save(buf)
        result['bytes_written'] = buf.tell()

        result['step'] = '5_upload'
        buf.seek(0)
        upload_result = handler.gdrive.upload_file(handler.file_id, buf)
        result['upload_response'] = str(upload_result)

        result['step'] = 'done'
        handler.close()
        return jsonify({'success': True, 'debug': result})

    except Exception as e:
        result['error'] = traceback.format_exc()
        return jsonify({'success': False, 'debug': result, 'error': str(e)}), 500


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

if __name__ == '__main__':
    load_deleted_workers()
    print("=" * 60)
    print("Machine Schedule Manager — Google Drive Edition")
    print("=" * 60)
    print(f"Matrix  : {MATRIX_FILE_ID or MATRIX_FILENAME}")
    print()
    print("Testing Google Drive connection…")
    try:
        get_excel_handler()
        print("✓ Google Drive connection successful")
    except FileNotFoundError as e:
        print(f"✗ File not found: {e}")
        print("  Run: python create_valid_template.py")
        raise SystemExit(1)
    except Exception as e:
        print(f"✗ Connection error: {e}")
        print("  Check credentials.json / token.pickle, then run: python test_gdrive_connection.py")
        raise SystemExit(1)

    print(f"\nServer starting at http://{HOST}:{PORT}")
    print("=" * 60)
    app.run(debug=DEBUG, host=HOST, port=PORT)