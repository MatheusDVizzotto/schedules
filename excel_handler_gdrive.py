"""
excel_handler_gdrive.py — Load/save the schedule spreadsheet via Google Drive.

The workbook layout assumed:
  - Column A            : machine names (in the row ranges defined in config)
  - Row 1               : worker names (in the column range defined in config)
  - Intersection cells  : proficiency codes (e.g. "1", "2", "C", "T", etc.)
  - Yellow-filled cells : schedule assignments written by this app
"""
import io
import re
from collections import defaultdict
from datetime import datetime, timedelta

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config import (
    CLEAR_PREVIOUS_SCHEDULE,
    MACHINE_RANGES,
    SCHEDULE_END_TIME,
    SCHEDULE_INTERVAL_MINUTES,
    SCHEDULE_MODE,
    SCHEDULE_START_TIME,
    WORKER_COL_END,
    WORKER_COL_START,
)
from google_drive_handler import GoogleDriveHandler

YELLOW = "FFFF00"
HEADER_BLUE = "CCE5FF"

ROLE_ABBREV = {
    'Main Role': 'MR',
    'Competent': 'C',
    'Trainee': 'T',
}


class ExcelHandlerGDrive:
    """Read/write the master schedule spreadsheet stored in Google Drive."""

    def __init__(self, file_id: str | None = None, filename: str | None = None):
        self.gdrive = GoogleDriveHandler()
        self.file_id = file_id
        self.filename = filename
        self.workbook: openpyxl.Workbook | None = None

        # Shared cell styles
        self.yellow_fill = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type="solid")
        self.header_fill = PatternFill(start_color=HEADER_BLUE, end_color=HEADER_BLUE, fill_type="solid")
        self.thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )
        self.center_align = Alignment(horizontal='center', vertical='center')
        self.left_align   = Alignment(horizontal='left',   vertical='center', wrap_text=True)

        # Resolve filename → file_id at construction time
        if not self.file_id and self.filename:
            self.file_id = self.gdrive.get_file_id_by_name(self.filename)
            if not self.file_id:
                raise FileNotFoundError(f"'{self.filename}' not found in Google Drive")

    # ------------------------------------------------------------------
    # Core I/O
    # ------------------------------------------------------------------

    def load(self):
        """Download and parse the workbook from Google Drive."""
        if not self.file_id:
            raise ValueError("No file_id available — check config.py")

        buf = self.gdrive.download_file(self.file_id)
        self.workbook = openpyxl.load_workbook(buf, data_only=True, keep_vba=False)
        return self.workbook

    def save(self):
        """Serialise and re-upload the workbook to Google Drive."""
        if not self.workbook:
            raise RuntimeError("No workbook loaded — call load() first")
        if not self.file_id:
            raise ValueError("No file_id available — check config.py")

        buf = io.BytesIO()
        self.workbook.save(buf)
        self.gdrive.upload_file(self.file_id, buf)

    def close(self):
        if self.workbook:
            self.workbook.close()
            self.workbook = None

    # ------------------------------------------------------------------
    # Sheet helpers
    # ------------------------------------------------------------------

    def _active_sheet(self, sheet_name: str | None = None):
        """Return the named sheet, or the active sheet if name is None."""
        if not self.workbook:
            raise RuntimeError("Workbook not loaded")
        if sheet_name:
            try:
                return self.workbook[sheet_name]
            except KeyError:
                pass
        return self.workbook.active

    # ------------------------------------------------------------------
    # Machines
    # ------------------------------------------------------------------

    def get_machines(self, sheet_name: str | None = None) -> list[dict]:
        """
        Return [{row, name}, …] for every non-empty cell in column A
        within the configured row ranges.
        """
        sheet = self._active_sheet(sheet_name)
        machines = []
        for start, end in MACHINE_RANGES:
            for row in range(start, end + 1):
                val = sheet[f'A{row}'].value
                if val:
                    machines.append({'row': row, 'name': str(val).strip()})
        return machines

    # ------------------------------------------------------------------
    # Workers
    # ------------------------------------------------------------------

    def get_workers(self, sheet_name: str | None = None) -> list[dict]:
        """
        Return [{col, col_letter, name}, …] for every non-empty cell in
        row 1 starting from WORKER_COL_START. Scans until 5 consecutive
        empty columns are found so newly added workers are always picked up
        without needing to restart the app or update WORKER_COL_END.
        """
        sheet = self._active_sheet(sheet_name)
        workers = []
        consecutive_empty = 0
        col = WORKER_COL_START
        while consecutive_empty < 5:
            letter = get_column_letter(col)
            val = sheet[f'{letter}1'].value
            if val:
                workers.append({'col': col, 'col_letter': letter, 'name': str(val).strip()})
                consecutive_empty = 0
            else:
                consecutive_empty += 1
            col += 1
        return workers

    def update_worker_name(self, worker_col: int, new_name: str, sheet_name: str | None = None):
        """Rename a worker in the header row and save."""
        sheet = self._active_sheet(sheet_name)
        sheet[f'{get_column_letter(worker_col)}1'] = new_name
        self.save()

    def delete_worker_permanently(self, worker_col: int, sheet_name: str | None = None):
        """Delete the worker's entire column from the spreadsheet, shifting subsequent columns left."""
        sheet = self._active_sheet(sheet_name)
        sheet.delete_cols(worker_col, 1)
        self.save()

    def add_worker(self, name: str, sheet_name: str | None = None) -> int:
        """
        Write *name* into the next empty column in row 1 after WORKER_COL_END.

        Styling applied:
          Header cell  — light orange fill, medium solid border on all sides, bold font
          Proficiency cells (one per machine row) — white fill, dotted borders between
            cells, medium solid border around the whole column as a group

        Returns the column number used.
        """
        sheet    = self._active_sheet(sheet_name)
        machines = self.get_machines(sheet_name)
        existing = self.get_workers(sheet_name)

        next_col = max(w['col'] for w in existing) + 1 if existing else WORKER_COL_START
        letter   = get_column_letter(next_col)

        # ── Shared style objects ───────────────────────────────────────────
        ORANGE      = 'FCE4D6'   # light orange (Excel "Orange, Accent 2, Lighter 80%")
        WHITE       = 'FFFFFF'
        MEDIUM      = 'medium'
        DOTTED      = 'dotted'

        orange_fill = PatternFill(start_color=ORANGE, end_color=ORANGE, fill_type='solid')
        white_fill  = PatternFill(start_color=WHITE,  end_color=WHITE,  fill_type='solid')
        bold_font   = openpyxl.styles.Font(bold=True)

        medium_side = Side(style=MEDIUM)
        dotted_side = Side(style=DOTTED)

        # Medium border on all four sides (used for header and outer proficiency cells)
        full_medium = Border(
            left=medium_side, right=medium_side,
            top=medium_side,  bottom=medium_side,
        )

        # ── Header cell (row 1) ───────────────────────────────────────────
        header = sheet[f'{letter}1']
        header.value     = name
        header.fill      = orange_fill
        header.font      = bold_font
        header.border    = full_medium
        header.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # ── Proficiency cells — every row from first to last (including separators) ──
        ranges = MACHINE_RANGES
        if not ranges:
            self.save()
            return next_col

        first_row = min(r[0] for r in ranges)
        last_row  = max(r[1] for r in ranges)
        all_rows  = list(range(first_row, last_row + 1))

        thin_side   = Side(style='thin')
        dashed_side = Side(style='dashed')

        for row in all_rows:
            cell = sheet[f'{letter}{row}']
            cell.value     = ''
            cell.fill      = white_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

            is_first = (row == first_row)
            is_last  = (row == last_row)

            cell.border = Border(
                left=thin_side,
                right=thin_side,
                top=medium_side   if is_first else dashed_side,
                bottom=thin_side  if is_last  else dashed_side,
            )

        print(f"  Added worker {name!r} at column {next_col} ({letter}), "
              f"styled {len(all_rows)} cells (rows {first_row}–{last_row})")
        self.save()
        return next_col

    # ------------------------------------------------------------------
    # Proficiency
    # ------------------------------------------------------------------

    def get_proficiency(
        self, machine_row: int, worker_col: int, sheet_name: str | None = None
    ) -> str:
        sheet = self._active_sheet(sheet_name)
        val = sheet[f'{get_column_letter(worker_col)}{machine_row}'].value
        return str(val).strip() if val is not None else ''

    def get_all_proficiencies(self, sheet_name: str | None = None) -> dict:
        """
        Return a nested dict:
          { machine_row: { worker_col: proficiency_value, … }, … }
        Keys are ints (not strings) for easy lookup.
        """
        sheet = self._active_sheet(sheet_name)
        machines = self.get_machines(sheet_name)
        workers = self.get_workers(sheet_name)

        result: dict[int, dict[int, str]] = {}
        for m in machines:
            result[m['row']] = {}
            for w in workers:
                val = sheet[f'{get_column_letter(w["col"])}{m["row"]}'].value
                result[m['row']][w['col']] = str(val).strip() if val is not None else ''
        return result

    # Canonical display names and fills for each proficiency level
    PROF_DISPLAY = {
        'MR': 'Main Role',
        'C':  'Competent',
        'T':  'Trainee',
        '':   '',
    }
    PROF_FILL = {
        'MR': PatternFill(start_color='00B050', end_color='00B050', fill_type='solid'),  # green
        'C':  PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid'),  # yellow
        'T':  PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid'),  # orange
        '':   PatternFill(fill_type=None),                                                # white/none
    }
    PROF_FONT = {
        'MR': openpyxl.styles.Font(bold=True, color='000000'),  # black
        'C':  openpyxl.styles.Font(bold=True, color='000000'),  # black
        'T':  openpyxl.styles.Font(bold=True, color='000000'),  # black
        '':   openpyxl.styles.Font(),
    }

    def update_proficiencies_bulk(self, proficiencies: dict, sheet_name: str | None = None):
        """
        Write multiple proficiency values and save once.

        Writes the full role name (e.g. "Main Role") into the cell, with
        colour-coded fill: green = Main Role, yellow = Competent, orange = Trainee.

        *proficiencies* format (keys may be int or str):
          { machine_row: { worker_col: canonical_value, … }, … }
          canonical_value: 'MR' | 'C' | 'T' | ''
        """
        sheet = self._active_sheet(sheet_name)
        for machine_row, worker_data in proficiencies.items():
            for worker_col, value in worker_data.items():
                cell = sheet[f'{get_column_letter(int(worker_col))}{int(machine_row)}']
                # Write full name; fall back to raw value if unrecognised
                cell.value = self.PROF_DISPLAY.get(value, value)
                cell.fill  = self.PROF_FILL.get(value, PatternFill(fill_type=None))
                cell.font  = self.PROF_FONT.get(value, openpyxl.styles.Font())
                cell.alignment = Alignment(horizontal='center', vertical='center')
        self.save()

    # ------------------------------------------------------------------
    # Schedule – save
    # ------------------------------------------------------------------

    def save_schedule_visual(self, date, schedule_data: list[dict]):
        """
        Write assignments to the workbook and upload to Google Drive.

        Supports two modes (set in config.py):
          SAME_SHEET       — writes yellow cells into the active master sheet
          NEW_SHEET_PER_DAY — creates/overwrites a tab named YYYY-MM-DD

        Prints detailed debug output to the server console so you can see
        exactly what is matched and written.
        """
        sep = "=" * 55
        print(f"\n{sep}")
        print(f"SAVE SCHEDULE  date={date}  mode={SCHEDULE_MODE}")
        print(f"  Entries received: {len(schedule_data)}")
        for i, e in enumerate(schedule_data):
            print(f"  [{i}] machine={e.get('machine')!r}  worker={e.get('worker')!r}"
                  f"  {e.get('time_start')}-{e.get('time_finish')}")
        print(sep)

        # ── Choose / create target sheet ──────────────────────────────────
        if SCHEDULE_MODE == 'NEW_SHEET_PER_DAY':
            sheet_name = date.strftime('%d-%m-%y') if hasattr(date, 'strftime') else str(date)
            if sheet_name in self.workbook.sheetnames:
                # Remove the sheet entirely so merged cells, column widths, and
                # all formatting are gone before we write fresh data.
                idx = self.workbook.sheetnames.index(sheet_name)
                self.workbook.remove(self.workbook[sheet_name])
                sheet = self.workbook.create_sheet(sheet_name, idx)
                print(f"  Cleared and recreated sheet: {sheet_name!r}")
            else:
                sheet = self.workbook.create_sheet(sheet_name)
                print(f"  Creating new sheet: {sheet_name!r}")
            self._write_day_sheet(sheet, date, schedule_data)
        else:
            # SAME_SHEET — write yellow cells into the master sheet
            sheet = self._active_sheet()
            print(f"  Writing into master sheet: {sheet.title!r}")
            if CLEAR_PREVIOUS_SCHEDULE:
                self._clear_yellow_cells(sheet)
            self._write_yellow_cells(sheet, schedule_data)

        print("  Uploading to Google Drive…")
        self.save()
        print("  ✓ Done")
        print(sep + "\n")

    # ── Internal write helpers ────────────────────────────────────────────

    def _write_yellow_cells(self, sheet, schedule_data: list[dict]):
        """Write assignments as yellow cells into an existing sheet layout."""
        machines = self.get_machines()
        workers  = self.get_workers()

        machine_map = {m['name']: m for m in machines}
        worker_map  = {w['name']: w for w in workers}

        print(f"  Machines in sheet ({len(machines)}): {[m['name'] for m in machines[:5]]}{'...' if len(machines)>5 else ''}")
        print(f"  Workers in sheet  ({len(workers)}):  {[w['name'] for w in workers[:5]]}{'...' if len(workers)>5 else ''}")

        saved   = 0
        skipped = []

        for entry in schedule_data:
            machine_name = str(entry.get('machine', '')).strip()
            worker_name  = str(entry.get('worker',  '')).strip()
            role         = entry.get('role', '')
            time_start   = entry.get('time_start', '')
            time_finish  = entry.get('time_finish', '')

            machine = machine_map.get(machine_name)
            worker  = worker_map.get(worker_name)

            if not machine:
                msg = f"machine {machine_name!r} not in sheet (check MACHINE_RANGES in config.py)"
                print(f"  ⚠ {msg}")
                skipped.append(msg)
                continue
            if not worker:
                msg = f"worker {worker_name!r} not in sheet (check WORKER_COL_START/END in config.py)"
                print(f"  ⚠ {msg}")
                skipped.append(msg)
                continue

            col_letter = get_column_letter(worker['col'])
            cell_ref   = f"{col_letter}{machine['row']}"
            cell       = sheet[cell_ref]

            abbrev = ROLE_ABBREV.get(role, '')
            notes  = str(entry.get('notes', '') or '').strip()
            text   = f"{time_start}-{time_finish}"
            if abbrev:
                text += f" ({abbrev})"
            if notes:
                text += f" | {notes}"

            cell.value     = text
            cell.fill      = self.yellow_fill
            cell.alignment = self.center_align
            print(f"  ✓ {cell_ref}  {machine_name!r} / {worker_name!r}  → {text!r}")
            saved += 1

        print(f"  Written: {saved}  Skipped: {len(skipped)}")

    def _write_day_sheet(self, sheet, date, schedule_data: list[dict]):
        """
        Write a visual day-schedule sheet that exactly mirrors the Main sheet:

        - Column A copies machine names in the same row order as Main
        - Blank separator rows between blocks are preserved (with their colours)
        - Each machine row copies its fill colour from the Main sheet
        - Columns B+ are time slots (06:00 … 22:00 in 30-min steps)
        - Unassigned machine rows: time cells are white/empty
        - Assigned rows: time cells for [start, finish) are yellow with worker name
        - Multiple workers on one machine each get their own sub-row
        """
        sheet_name = date.strftime('%d-%m-%y') if hasattr(date, 'strftime') else str(date)
        date_str   = date.strftime('%d/%m/%y') if hasattr(date, 'strftime') else str(date)
        main_sheet = self._active_sheet()           # read colours from Main
        time_slots = self._generate_time_slots()    # ['06:00', '06:30', …, '22:00']
        no_fill   = PatternFill(fill_type=None)
        bold_font = openpyxl.styles.Font(bold=True)

        # ── Build assignment lookup: machine_name → merged single entry ────
        # All workers on the same machine share one row.
        # Per-worker times are encoded in the cell text: "Alice 07:00-15:00, Bob 08:00-16:00"
        # Yellow block spans earliest start → latest finish across all workers.
        machine_entries = {}   # mname → {worker_entries, notes, time_start, time_finish}
        for entry in schedule_data:
            mname      = str(entry.get('machine', '')).strip()
            worker     = str(entry.get('worker',  '')).strip()
            note       = str(entry.get('notes',   '') or '').strip()
            t_start    = entry.get('time_start',  '')
            t_finish   = entry.get('time_finish', '')

            if mname not in machine_entries:
                machine_entries[mname] = {
                    'worker_entries': [],
                    'notes':          [],
                    'time_start':     t_start,
                    'time_finish':    t_finish,
                }

            e = machine_entries[mname]
            if worker:
                e['worker_entries'].append({'name': worker, 'time_start': t_start, 'time_finish': t_finish})
            if note and note not in e['notes']:
                e['notes'].append(note)

            # Expand merged time range to cover all workers
            if t_start and (not e['time_start'] or t_start < e['time_start']):
                e['time_start'] = t_start
            if t_finish and (not e['time_finish'] or t_finish > e['time_finish']):
                e['time_finish'] = t_finish

        # ── Col layout ────────────────────────────────────────────────────
        # Col A  = machine name
        # Col B+ = time slots  (no separate notes column)
        T_OFF = 2   # time slots start at col B

        # ── Row 1: title cell + time-slot headers ─────────────────────────
        title = sheet.cell(row=1, column=1, value=f'Schedule — {date_str}')  # DD/MM/YY display
        title.font      = openpyxl.styles.Font(bold=True, size=12)
        title.fill      = self.header_fill
        title.alignment = self.center_align

        for i, slot in enumerate(time_slots):
            c = sheet.cell(row=1, column=T_OFF + i, value=slot)
            c.fill      = self.header_fill
            c.font      = bold_font
            c.alignment = self.center_align
            c.border    = self.thin_border
            sheet.column_dimensions[get_column_letter(T_OFF + i)].width = 7

        sheet.column_dimensions['A'].width = 32
        sheet.row_dimensions[1].height = 20

        # ── Build ordered row list (mirrors Main sheet row order) ──────────
        rows_to_render = []   # [(main_row, kind)]  kind = 'machine'|'separator'
        ranges   = MACHINE_RANGES
        all_rows = set()
        for start, end in ranges:
            for r in range(start, end + 1):
                all_rows.add(r)

        if ranges:
            first_row = min(r[0] for r in ranges)
            last_row  = max(r[1] for r in ranges)
            for main_row in range(first_row, last_row + 1):
                kind = 'machine' if main_row in all_rows else 'separator'
                rows_to_render.append((main_row, kind))

        # ── Write one row per machine (or separator) ──────────────────────
        dest_row = 2

        for main_row, kind in rows_to_render:

            # Copy row fill colour from Main sheet column A
            main_cell_a = main_sheet.cell(row=main_row, column=1)
            src_fill    = main_cell_a.fill
            if src_fill and src_fill.fill_type == 'solid' and src_fill.fgColor:
                rgb      = src_fill.fgColor.rgb
                row_fill = PatternFill(start_color=rgb, end_color=rgb, fill_type='solid')
            else:
                row_fill = no_fill

            if kind == 'separator':
                sheet.cell(row=dest_row, column=1, value='').fill = row_fill
                for i in range(len(time_slots)):
                    sheet.cell(row=dest_row, column=T_OFF + i, value='').fill = row_fill
                sheet.row_dimensions[dest_row].height =                     main_sheet.row_dimensions[main_row].height or 8
                dest_row += 1
                continue

            # ── Machine row: always exactly ONE row ────────────────────────
            mname = str(main_cell_a.value or '').strip()
            entry = machine_entries.get(mname)   # None if unassigned

            # Col A — machine name
            a_cell = sheet.cell(row=dest_row, column=1, value=mname)
            a_cell.fill      = row_fill
            a_cell.font      = openpyxl.styles.Font(
                bold=bool(main_cell_a.font and main_cell_a.font.bold))
            a_cell.alignment = Alignment(vertical='center', wrap_text=True)
            a_cell.border    = self.thin_border

            # Time slots — white/empty by default
            for i in range(len(time_slots)):
                c = sheet.cell(row=dest_row, column=T_OFF + i, value='')
                c.fill   = no_fill
                c.border = self.thin_border

            # Yellow bar strategy:
            # 1. Merge overlapping worker intervals into contiguous bands.  Workers
            #    whose times overlap (or are bridged by another worker) share ONE
            #    yellow merged cell.  Non-overlapping blocks (genuine gap = no one
            #    working) produce separate yellow cells at their correct positions.
            # 2. Yellow fill is applied per-band, so a gap between two bands stays
            #    white — the bar does NOT stretch across empty time.
            # 3. Text format:  single-block worker → "Name HH:MM-HH:MM"
            #                  multi-block worker  → "Name (HH:MM-HH:MM),(HH:MM-HH:MM)"
            #    Single-block workers listed first, then multi-block workers.
            # 4. Notes appear only on the first (leftmost) band.
            if entry:
                notes_text = ' | '.join(entry['notes']) if entry['notes'] else ''

                all_intervals = sorted(
                    [(we['time_start'], we['time_finish'], we['name'])
                     for we in entry['worker_entries']
                     if we['time_start'] and we['time_finish']],
                    key=lambda x: x[0],
                )

                if all_intervals:
                    bands: list[tuple] = []  # [(band_start, band_finish, [(name, ws, wf), ...])]
                    for w_start, w_finish, name in all_intervals:
                        if bands and w_start <= bands[-1][1]:
                            b_start, b_finish, workers = bands[-1]
                            bands[-1] = (b_start, max(b_finish, w_finish),
                                         workers + [(name, w_start, w_finish)])
                        else:
                            bands.append((w_start, w_finish, [(name, w_start, w_finish)]))

                    first_band = True
                    for band_start, band_finish, band_workers in bands:
                        # Group by worker name, preserving each worker's block order
                        worker_blocks: dict[str, list[tuple[str, str]]] = {}
                        for name, ws, wf in band_workers:
                            worker_blocks.setdefault(name, []).append((ws, wf))

                        # Single-block workers first (sorted by start then name),
                        # then multi-block workers in the same order
                        single = [(n, b) for n, b in worker_blocks.items() if len(b) == 1]
                        multi  = [(n, b) for n, b in worker_blocks.items() if len(b) > 1]
                        single.sort(key=lambda x: (x[1][0][0], x[0]))
                        multi.sort(key=lambda x: (x[1][0][0], x[0]))

                        worker_parts = []
                        for name, blocks in single + multi:
                            if len(blocks) == 1:
                                ws, wf = blocks[0]
                                worker_parts.append(f"{name} {ws}-{wf}")
                            else:
                                block_strs = ','.join(f"({ws}-{wf})" for ws, wf in blocks)
                                worker_parts.append(f"{name} {block_strs}")

                        cell_text = ', '.join(worker_parts)
                        if first_band and notes_text:
                            cell_text += f' - {notes_text}'
                        first_band = False

                        start_idx  = self._slot_index(band_start, time_slots)
                        finish_idx = self._slot_index(band_finish, time_slots)
                        if start_idx is None or finish_idx is None or start_idx >= finish_idx:
                            print(f"  ⚠ Bad times for {mname!r} ({band_start}–{band_finish})")
                            continue

                        col_start = T_OFF + start_idx
                        col_end   = T_OFF + min(finish_idx, len(time_slots)) - 1

                        for idx in range(start_idx, min(finish_idx, len(time_slots))):
                            c = sheet.cell(row=dest_row, column=T_OFF + idx)
                            c.fill   = self.yellow_fill
                            c.border = self.thin_border

                        if col_end > col_start:
                            try:
                                sheet.merge_cells(
                                    start_row=dest_row, start_column=col_start,
                                    end_row=dest_row,   end_column=col_end,
                                )
                            except Exception:
                                pass

                        merged_cell           = sheet.cell(row=dest_row, column=col_start)
                        merged_cell.value     = cell_text
                        merged_cell.fill      = self.yellow_fill
                        merged_cell.alignment = self.left_align
                        merged_cell.border    = self.thin_border
                        print(f"  ✓ {mname!r}  {band_start}→{band_finish}  {cell_text!r}")

            sheet.row_dimensions[dest_row].height =                 main_sheet.row_dimensions[main_row].height or 15
            dest_row += 1

        print(f"  ✓ Day sheet {date_str!r}: "
              f"{len(rows_to_render)} rows mirrored, "
              f"{len(machine_entries)} machines assigned")

    def _slot_index(self, time_str: str, time_slots: list[str]) -> int | None:
        """Return the index in time_slots for time_str, rounding to nearest slot."""
        if not time_str:
            return None
        try:
            t = datetime.strptime(time_str, '%H:%M')
        except ValueError:
            return None
        for i, slot in enumerate(time_slots):
            if datetime.strptime(slot, '%H:%M') >= t:
                return i
        return len(time_slots)   # past the end → clamp to last slot

    def _clear_yellow_cells(self, sheet):
        """Remove value and yellow fill from any yellow-filled cell."""
        no_fill = PatternFill(fill_type=None)
        cleared = 0
        for row in sheet.iter_rows():
            for cell in row:
                if (
                    cell.fill
                    and cell.fill.fill_type == 'solid'
                    and cell.fill.fgColor
                    and cell.fill.fgColor.rgb == f'FF{YELLOW}'
                ):
                    cell.value = None
                    cell.fill = no_fill
                    cleared += 1
        print(f"  Cleared {cleared} yellow cells")

    # ------------------------------------------------------------------
    # Schedule – read
    # ------------------------------------------------------------------

    def get_schedule(self, date) -> list[dict]:
        """
        Read back a saved day schedule from its dedicated sheet tab.

        For NEW_SHEET_PER_DAY mode the tab is named YYYY-MM-DD.
        Parses each yellow cell's text back into structured fields:
            machine, worker(s), time_start, time_finish, notes

        Cell text format written by _write_day_sheet:
            "Worker1, Worker2 - notes"   (with notes)
            "Worker1, Worker2"           (no notes)

        The first yellow cell in each machine row holds the text;
        we find time_start/finish by scanning which slots are yellow.
        """
        sheet_name = date.strftime('%d-%m-%y') if hasattr(date, 'strftime') else str(date)

        if sheet_name not in self.workbook.sheetnames:
            print(f"  No sheet named {sheet_name!r} found. Available: {self.workbook.sheetnames}")
            return []

        sheet     = self.workbook[sheet_name]
        time_slots = self._generate_time_slots()
        T_OFF      = 2   # time slots start at column B (col index 2)

        # Build machine list from row 2 onwards in the day sheet
        # (row 1 is the header)
        schedule = []

        for row in sheet.iter_rows(min_row=2):
            machine_cell = row[0]   # Column A
            mname = str(machine_cell.value or '').strip()
            if not mname:
                continue   # separator row

            # Find the assigned time range.
            #
            # Strategy: find the cell with a non-empty text value — that is always
            # the first yellow slot and holds "Worker - notes".
            # Then scan outward to find the contiguous block of yellow-filled cells.
            # We identify yellow by RGB; if the RGB is unreliable after a Drive
            # round-trip we fall back to checking the fill type and comparing against
            # the known non-yellow fills (the block colours on col A are NOT applied
            # to time-slot cells — those are explicitly set to PatternFill(fill_type=None)).
            #
            # In _write_day_sheet:
            #   - Unassigned time cells: fill_type=None  (transparent/white)
            #   - Assigned time cells:   fill_type='solid', fgColor=FFFF00
            # So fill_type == 'solid' on a time-slot cell means it is yellow.

            time_cells = list(row[1:])   # skip col A

            # Collect ALL text cells — each is the top-left of a separate merged block
            text_blocks = [
                (col_idx, str(cell.value).strip())
                for col_idx, cell in enumerate(time_cells)
                if cell.value is not None and str(cell.value).strip()
            ]

            if not text_blocks:
                continue   # no text → unassigned row

            # Patterns for the three cell text formats:
            #   plain   "Worker A 07:00-15:00"
            #   paren1  "Worker B (07:00-10:00)"   ← first block of a multi-block worker
            #   paren+  "(12:00-15:00)"             ← continuation block for previous worker
            _re_plain  = re.compile(r'^(.*?)\s+(\d{2}:\d{2})-(\d{2}:\d{2})$')
            _re_paren1 = re.compile(r'^(.*?)\s+\((\d{2}:\d{2})-(\d{2}:\d{2})\)$')
            _re_paren  = re.compile(r'^\((\d{2}:\d{2})-(\d{2}:\d{2})\)$')

            for text_col, block_text in text_blocks:
                print(f"  Row {mname!r}: slot {text_col}, text={block_text!r}")

                # Parse notes
                workers_str = block_text
                notes       = ''
                if ' - ' in block_text:
                    parts       = block_text.split(' - ', 1)
                    workers_str = parts[0].strip()
                    notes       = parts[1].strip()

                current_worker = None
                for part in workers_str.split(','):
                    part = part.strip()
                    if not part:
                        continue

                    # "(HH:MM-HH:MM)" — extra block for the most recent worker
                    mp = _re_paren.match(part)
                    if mp and current_worker:
                        wname    = current_worker
                        w_start  = mp.group(1)
                        w_finish = mp.group(2)
                    else:
                        # "Name (HH:MM-HH:MM)" — first paren block for this worker
                        mp1 = _re_paren1.match(part)
                        if mp1:
                            current_worker = mp1.group(1).strip()
                            wname    = current_worker
                            w_start  = mp1.group(2)
                            w_finish = mp1.group(3)
                        else:
                            # "Name HH:MM-HH:MM" — plain single block
                            mpl = _re_plain.match(part)
                            if mpl:
                                current_worker = mpl.group(1).strip()
                                wname    = current_worker
                                w_start  = mpl.group(2)
                                w_finish = mpl.group(3)
                            else:
                                # Legacy: no times — use slot position as fallback
                                current_worker = part
                                wname    = part
                                w_start  = time_slots[text_col] if text_col < len(time_slots) else ''
                                w_finish = ''
                    schedule.append({
                        'machine':     mname,
                        'worker':      wname,
                        'time_start':  w_start,
                        'time_finish': w_finish,
                        'notes':       notes,
                    })

        print(f"  Loaded {len(schedule)} assignments from sheet {sheet_name!r}")
        return schedule

    # ------------------------------------------------------------------
    # Time utilities (kept for potential future use)
    # ------------------------------------------------------------------

    def _generate_time_slots(
        self,
        start: str = SCHEDULE_START_TIME,
        end: str = SCHEDULE_END_TIME,
        interval: int = SCHEDULE_INTERVAL_MINUTES,
    ) -> list[str]:
        slots = []
        current = datetime.strptime(start, "%H:%M")
        stop = datetime.strptime(end, "%H:%M")
        while current <= stop:
            slots.append(current.strftime("%H:%M"))
            current += timedelta(minutes=interval)
        return slots