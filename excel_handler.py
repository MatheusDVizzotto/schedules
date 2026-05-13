import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from datetime import datetime, timedelta
import os

class ExcelHandler:
    def __init__(self, file_path):
        self.file_path = file_path
        self.workbook = None
        self.yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        self.header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def load(self):
        """Load the Excel workbook"""
        if os.path.exists(self.file_path):
            self.workbook = load_workbook(self.file_path)
        else:
            self.workbook = Workbook()
        return self.workbook

    def get_machines(self, sheet_name=None):
        """Extract machines from specified ranges: A5:A8, A10:A22, A24:A36"""
        if not self.workbook:
            self.load()

        sheet = self.workbook[sheet_name] if sheet_name else self.workbook.active
        machines = []

        # Define machine ranges
        ranges = [
            (5, 8),   # A5:A8
            (10, 22), # A10:A22
            (24, 36)  # A24:A36
        ]

        for start, end in ranges:
            for row in range(start, end + 1):
                cell_value = sheet[f'A{row}'].value
                if cell_value:
                    machines.append({
                        'row': row,
                        'name': str(cell_value).strip()
                    })

        return machines

    def get_workers(self, sheet_name=None):
        """Extract workers from row 1 starting at column AO (41), scanning until 5 consecutive empty columns."""
        if not self.workbook:
            self.load()

        sheet = self.workbook[sheet_name] if sheet_name else self.workbook.active
        workers = []

        consecutive_empty = 0
        col = 41  # AO
        while consecutive_empty < 5:
            col_letter = get_column_letter(col)
            cell_value = sheet[f'{col_letter}1'].value
            if cell_value:
                workers.append({
                    'col': col,
                    'col_letter': col_letter,
                    'name': str(cell_value).strip()
                })
                consecutive_empty = 0
            else:
                consecutive_empty += 1
            col += 1

        return workers

    def get_proficiency(self, machine_row, worker_col, sheet_name=None):
        """Get proficiency level for a machine-worker combination"""
        if not self.workbook:
            self.load()

        sheet = self.workbook[sheet_name] if sheet_name else self.workbook.active

        # Get the cell at intersection
        col_letter = get_column_letter(worker_col)
        cell_value = sheet[f'{col_letter}{machine_row}'].value

        return cell_value if cell_value else ''

    def get_all_proficiencies(self, sheet_name=None):
        """Get all proficiency data as a matrix"""
        if not self.workbook:
            self.load()

        sheet = self.workbook[sheet_name] if sheet_name else self.workbook.active
        machines = self.get_machines(sheet_name)
        workers = self.get_workers(sheet_name)

        proficiencies = {}
        for machine in machines:
            proficiencies[machine['row']] = {}
            for worker in workers:
                col_letter = get_column_letter(worker['col'])
                cell_value = sheet[f'{col_letter}{machine["row"]}'].value
                proficiencies[machine['row']][worker['col']] = cell_value if cell_value else ''

        return proficiencies

    def update_worker_name(self, worker_col, new_name, sheet_name=None):
        """Update worker name in the header row"""
        if not self.workbook:
            self.load()

        sheet = self.workbook[sheet_name] if sheet_name else self.workbook.active
        col_letter = get_column_letter(worker_col)
        sheet[f'{col_letter}1'] = new_name
        self.workbook.save(self.file_path)

    def update_proficiency(self, machine_row, worker_col, proficiency, sheet_name=None):
        """Update proficiency level for a machine-worker combination"""
        if not self.workbook:
            self.load()

        sheet = self.workbook[sheet_name] if sheet_name else self.workbook.active
        col_letter = get_column_letter(worker_col)
        sheet[f'{col_letter}{machine_row}'] = proficiency
        self.workbook.save(self.file_path)

    def update_proficiencies_bulk(self, proficiencies, sheet_name=None):
        """Update multiple proficiencies at once"""
        if not self.workbook:
            self.load()

        sheet = self.workbook[sheet_name] if sheet_name else self.workbook.active

        for machine_row, worker_data in proficiencies.items():
            for worker_col, proficiency in worker_data.items():
                col_letter = get_column_letter(worker_col)
                sheet[f'{col_letter}{machine_row}'] = proficiency

        self.workbook.save(self.file_path)

    def get_or_create_day_sheet(self, date):
        """Get or create a sheet for specific date"""
        if not self.workbook:
            self.load()

        sheet_name = date.strftime('%Y-%m-%d')

        if sheet_name in self.workbook.sheetnames:
            return self.workbook[sheet_name]
        else:
            # Create new sheet
            sheet = self.workbook.create_sheet(sheet_name)
            return sheet

    def _generate_time_columns(self, start_time="06:00", end_time="22:00", interval_minutes=30):
        """Generate time slots for the day"""
        time_slots = []
        current = datetime.strptime(start_time, "%H:%M")
        end = datetime.strptime(end_time, "%H:%M")

        while current <= end:
            time_slots.append(current.strftime("%H:%M"))
            current += timedelta(minutes=interval_minutes)

        return time_slots

    def _time_to_column(self, time_str, time_slots, start_col=2):
        """Convert time string to column number"""
        if not time_str or time_str == "":
            return None

        try:
            # Normalize time format
            time_obj = datetime.strptime(time_str, "%H:%M")
            time_formatted = time_obj.strftime("%H:%M")

            if time_formatted in time_slots:
                return start_col + time_slots.index(time_formatted)

            # Find closest time slot
            for i, slot in enumerate(time_slots):
                slot_obj = datetime.strptime(slot, "%H:%M")
                if time_obj <= slot_obj:
                    return start_col + i

            return start_col + len(time_slots) - 1
        except:
            return None

    def save_schedule_visual(self, date, schedule_data):
        """Save schedule with visual formatting (yellow cells with worker names)"""
        sheet = self.get_or_create_day_sheet(date)

        # Clear sheet
        for row in sheet.iter_rows():
            for cell in row:
                cell.value = None
                cell.fill = PatternFill(fill_type=None)
                cell.border = None

        # Generate time slots (6:00 AM to 10:00 PM, 30-min intervals)
        time_slots = self._generate_time_columns("06:00", "22:00", 30)

        # Get machines
        machines = self.get_machines()

        # Setup headers
        sheet['A1'] = 'Machine / Time'
        sheet['A1'].fill = self.header_fill
        sheet['A1'].font = Font(bold=True)
        sheet['A1'].border = self.border

        # Time headers (starting from column B)
        for idx, time_slot in enumerate(time_slots):
            col = idx + 2  # Start from column B (2)
            col_letter = get_column_letter(col)
            sheet[f'{col_letter}1'] = time_slot
            sheet[f'{col_letter}1'].fill = self.header_fill
            sheet[f'{col_letter}1'].font = Font(bold=True)
            sheet[f'{col_letter}1'].border = self.border
            sheet[f'{col_letter}1'].alignment = Alignment(horizontal='center')
            sheet.column_dimensions[col_letter].width = 6

        # Set machine names in column A
        machine_row_map = {}
        for idx, machine in enumerate(machines):
            row = idx + 2  # Start from row 2
            sheet[f'A{row}'] = machine['name']
            sheet[f'A{row}'].fill = self.header_fill
            sheet[f'A{row}'].font = Font(bold=True)
            sheet[f'A{row}'].border = self.border
            machine_row_map[machine['name']] = row

        sheet.column_dimensions['A'].width = 25

        # Group schedule by machine
        schedule_by_machine = {}
        for entry in schedule_data:
            machine_name = entry.get('machine')
            if machine_name not in schedule_by_machine:
                schedule_by_machine[machine_name] = []
            schedule_by_machine[machine_name].append(entry)

        # Fill schedule with yellow cells and worker names
        for machine_name, entries in schedule_by_machine.items():
            if machine_name not in machine_row_map:
                continue

            row = machine_row_map[machine_name]

            for entry in entries:
                time_start = entry.get('time_start', '')
                time_finish = entry.get('time_finish', '')
                worker_name = entry.get('worker', '')
                role = entry.get('role', '')

                if not time_start or not time_finish:
                    continue

                # Get column indices for start and finish times
                start_col = self._time_to_column(time_start, time_slots)
                end_col = self._time_to_column(time_finish, time_slots)

                if start_col and end_col:
                    # Fill cells with yellow and add worker name
                    worker_text = f"{worker_name}"
                    if role:
                        role_abbrev = {'Main Role': 'MR', 'Competent': 'C', 'Trainee': 'T'}.get(role, '')
                        if role_abbrev:
                            worker_text += f" ({role_abbrev})"

                    for col in range(start_col, end_col + 1):
                        col_letter = get_column_letter(col)
                        cell = sheet[f'{col_letter}{row}']

                        # Add worker name to first cell
                        if col == start_col:
                            cell.value = worker_text

                        # Apply yellow fill
                        cell.fill = self.yellow_fill
                        cell.border = self.border
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.font = Font(size=9)

        self.workbook.save(self.file_path)

    def get_schedule(self, date):
        """Get schedule data for a specific date"""
        sheet_name = date.strftime('%Y-%m-%d')

        if not self.workbook:
            self.load()

        if sheet_name not in self.workbook.sheetnames:
            return []

        sheet = self.workbook[sheet_name]
        schedule = []

        return schedule

    def close(self):
        """Close the workbook"""
        if self.workbook:
            self.workbook.close()
