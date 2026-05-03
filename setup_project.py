# setup_project.py - Run this file to create all project files
import os

# Create directories
print("Creating directory structure...")
os.makedirs('templates', exist_ok=True)
os.makedirs('static/css', exist_ok=True)
os.makedirs('static/js', exist_ok=True)
print("✓ Directories created")

# File contents dictionary
files = {
    'requirements.txt': '''Flask==3.0.0
openpyxl==3.1.2
python-dateutil==2.8.2
Werkzeug==3.0.1
''',

    'deleted_workers.json': '[]',

    'excel_handler.py': '''import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from datetime import datetime, timedelta
import os

class ExcelHandler:
    def __init__(self, file_path):
        self.file_path = file_path
        self.workbook = None
        self.yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        self.header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def load(self):
        """Load the Excel workbook"""
        if os.path.exists(self.file_path):
            self.workbook = load_workbook(self.file_path)
        else:
            self.workbook = Workbook()
        return self.workbook

    def get_machines(self, sheet_name=None):
        """Extract machines from specified ranges: A5:A8, A10:A22, A24:A36"""
        if not self.workbook:
            self.load()

        sheet = self.workbook[sheet_name] if sheet_name else self.workbook.active
        machines = []

        # Define machine ranges
        ranges = [
            (5, 8),   # A5:A8
            (10, 22), # A10:A22
            (24, 36)  # A24:A36
        ]

        for start, end in ranges:
            for row in range(start, end + 1):
                cell_value = sheet[f'A{row}'].value
                if cell_value:
                    machines.append({
                        'row': row,
                        'name': str(cell_value).strip()
                    })

        return machines

    def get_workers(self, sheet_name=None):
        """Extract workers from AO1:AX1 (columns 41-50)"""
        if not self.workbook:
            self.load()

        sheet = self.workbook[sheet_name] if sheet_name else self.workbook.active
        workers = []

        # AO = column 41, AX = column 50
        for col in range(41, 51):
            col_letter = get_column_letter(col)
            cell_value = sheet[f'{col_letter}1'].value
            if cell_value:
                workers.append({
                    'col': col,
                    'col_letter': col_letter,
                    'name': str(cell_value).strip()
                })

        return workers

    def get_proficiency(self, machine_row, worker_col, sheet_name=None):
        """Get proficiency level for a machine-worker combination"""
        if not self.workbook:
            self.load()

        sheet = self.workbook[sheet_name] if sheet_name else self.workbook.active

        # Get the cell at intersection
        col_letter = get_column_letter(worker_col)
        cell_value = sheet[f'{col_letter}{machine_row}'].value

        return cell_value if cell_value else ''

    def get_all_proficiencies(self, sheet_name=None):
        """Get all proficiency data as a matrix"""
        if not self.workbook:
            self.load()

        sheet = self.workbook[sheet_name] if sheet_name else self.workbook.active
        machines = self.get_machines(sheet_name)
        workers = self.get_workers(sheet_name)

        proficiencies = {}
        for machine in machines:
            proficiencies[machine['row']] = {}
            for worker in workers:
                col_letter = get_column_letter(worker['col'])
                cell_value = sheet[f'{col_letter}{machine["row"]}'].value
                proficiencies[machine['row']][worker['col']] = cell_value if cell_value else ''

        return proficiencies

    def update_worker_name(self, worker_col, new_name, sheet_name=None):
        """Update worker name in the header row"""
        if not self.workbook:
            self.load()

        sheet = self.workbook[sheet_name] if sheet_name else self.workbook.active
        col_letter = get_column_letter(worker_col)
        sheet[f'{col_letter}1'] = new_name
        self.workbook.save(self.file_path)

    def update_proficiency(self, machine_row, worker_col, proficiency, sheet_name=None):
        """Update proficiency level for a machine-worker combination"""
        if not self.workbook:
            self.load()

        sheet = self.workbook[sheet_name] if sheet_name else self.workbook.active
        col_letter = get_column_letter(worker_col)
        sheet[f'{col_letter}{machine_row}'] = proficiency
        self.workbook.save(self.file_path)

    def update_proficiencies_bulk(self, proficiencies, sheet_name=None):
        """Update multiple proficiencies at once"""
        if not self.workbook:
            self.load()

        sheet = self.workbook[sheet_name] if sheet_name else self.workbook.active

        for machine_row, worker_data in proficiencies.items():
            for worker_col, proficiency in worker_data.items():
                col_letter = get_column_letter(worker_col)
                sheet[f'{col_letter}{machine_row}'] = proficiency

        self.workbook.save(self.file_path)

    def get_or_create_day_sheet(self, date):
        """Get or create a sheet for specific date"""
        if not self.workbook:
            self.load()

        sheet_name = date.strftime('%Y-%m-%d')

        if sheet_name in self.workbook.sheetnames:
            return self.workbook[sheet_name]
        else:
            # Create new sheet
            sheet = self.workbook.create_sheet(sheet_name)
            return sheet

    def _generate_time_columns(self, start_time="06:00", end_time="22:00", interval_minutes=30):
        """Generate time slots for the day"""
        time_slots = []
        current = datetime.strptime(start_time, "%H:%M")
        end = datetime.strptime(end_time, "%H:%M")

        while current <= end:
            time_slots.append(current.strftime("%H:%M"))
            current += timedelta(minutes=interval_minutes)

        return time_slots

    def _time_to_column(self, time_str, time_slots, start_col=2):
        """Convert time string to column number"""
        if not time_str or time_str == "":
            return None

        try:
            # Normalize time format
            time_obj = datetime.strptime(time_str, "%H:%M")
            time_formatted = time_obj.strftime("%H:%M")

            if time_formatted in time_slots:
                return start_col + time_slots.index(time_formatted)

            # Find closest time slot
            for i, slot in enumerate(time_slots):
                slot_obj = datetime.strptime(slot, "%H:%M")
                if time_obj <= slot_obj:
                    return start_col + i

            return start_col + len(time_slots) - 1
        except:
            return None

    def save_schedule_visual(self, date, schedule_data):
        """Save schedule with visual formatting (yellow cells with worker names)"""
        sheet = self.get_or_create_day_sheet(date)

        # Clear sheet
        for row in sheet.iter_rows():
            for cell in row:
                cell.value = None
                cell.fill = PatternFill(fill_type=None)
                cell.border = None

        # Generate time slots (6:00 AM to 10:00 PM, 30-min intervals)
        time_slots = self._generate_time_columns("06:00", "22:00", 30)

        # Get machines
        machines = self.get_machines()

        # Setup headers
        sheet['A1'] = 'Machine / Time'
        sheet['A1'].fill = self.header_fill
        sheet['A1'].font = Font(bold=True)
        sheet['A1'].border = self.border

        # Time headers (starting from column B)
        for idx, time_slot in enumerate(time_slots):
            col = idx + 2  # Start from column B (2)
            col_letter = get_column_letter(col)
            sheet[f'{col_letter}1'] = time_slot
            sheet[f'{col_letter}1'].fill = self.header_fill
            sheet[f'{col_letter}1'].font = Font(bold=True)
            sheet[f'{col_letter}1'].border = self.border
            sheet[f'{col_letter}1'].alignment = Alignment(horizontal='center')
            sheet.column_dimensions[col_letter].width = 6

        # Set machine names in column A
        machine_row_map = {}
        for idx, machine in enumerate(machines):
            row = idx + 2  # Start from row 2
            sheet[f'A{row}'] = machine['name']
            sheet[f'A{row}'].fill = self.header_fill
            sheet[f'A{row}'].font = Font(bold=True)
            sheet[f'A{row}'].border = self.border
            machine_row_map[machine['name']] = row

        sheet.column_dimensions['A'].width = 25

        # Group schedule by machine
        schedule_by_machine = {}
        for entry in schedule_data:
            machine_name = entry.get('machine')
            if machine_name not in schedule_by_machine:
                schedule_by_machine[machine_name] = []
            schedule_by_machine[machine_name].append(entry)

        # Fill schedule with yellow cells and worker names
        for machine_name, entries in schedule_by_machine.items():
            if machine_name not in machine_row_map:
                continue

            row = machine_row_map[machine_name]

            for entry in entries:
                time_start = entry.get('time_start', '')
                time_finish = entry.get('time_finish', '')
                worker_name = entry.get('worker', '')
                role = entry.get('role', '')

                if not time_start or not time_finish:
                    continue

                # Get column indices for start and finish times
                start_col = self._time_to_column(time_start, time_slots)
                end_col = self._time_to_column(time_finish, time_slots)

                if start_col and end_col:
                    # Fill cells with yellow and add worker name
                    worker_text = f"{worker_name}"
                    if role:
                        role_abbrev = {'Main Role': 'MR', 'Competent': 'C', 'Trainee': 'T'}.get(role, '')
                        if role_abbrev:
                            worker_text += f" ({role_abbrev})"

                    for col in range(start_col, end_col + 1):
                        col_letter = get_column_letter(col)
                        cell = sheet[f'{col_letter}{row}']

                        # Add worker name to first cell
                        if col == start_col:
                            cell.value = worker_text

                        # Apply yellow fill
                        cell.fill = self.yellow_fill
                        cell.border = self.border
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.font = Font(size=9)

        self.workbook.save(self.file_path)

    def get_schedule(self, date):
        """Get schedule data for a specific date"""
        sheet_name = date.strftime('%Y-%m-%d')

        if not self.workbook:
            self.load()

        if sheet_name not in self.workbook.sheetnames:
            return []

        sheet = self.workbook[sheet_name]
        schedule = []

        return schedule

    def close(self):
        """Close the workbook"""
        if self.workbook:
            self.workbook.close()
''',

    'app.py': '''from flask import Flask, render_template, request, jsonify
from datetime import datetime, date
from excel_handler import ExcelHandler
import json
import os

app = Flask(__name__)
app.config['SECRET_KEY'] = 'your-secret-key-here'

# Path to Excel file
EXCEL_FILE_PATH = r'C:\\Users\\mtheu\\Downloads\\test.xlsx'

# Store deleted workers (in production, use database)
deleted_workers = set()

def load_deleted_workers():
    """Load deleted workers from file"""
    global deleted_workers
    if os.path.exists('deleted_workers.json'):
        with open('deleted_workers.json', 'r') as f:
            deleted_workers = set(json.load(f))

def save_deleted_workers():
    """Save deleted workers to file"""
    with open('deleted_workers.json', 'w') as f:
        json.dump(list(deleted_workers), f)

@app.route('/')
def index():
    """Main scheduling interface"""
    return render_template('index.html')

@app.route('/admin')
def admin():
    """Admin interface for worker management"""
    return render_template('admin.html')

@app.route('/workers')
def workers_page():
    """Worker edit interface"""
    return render_template('workers.html')

@app.route('/api/machines', methods=['GET'])
def get_machines():
    """API endpoint to get all machines"""
    try:
        handler = ExcelHandler(EXCEL_FILE_PATH)
        handler.load()
        machines = handler.get_machines()
        handler.close()
        return jsonify({'success': True, 'machines': machines})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/workers', methods=['GET'])
def get_workers():
    """API endpoint to get all workers"""
    try:
        handler = ExcelHandler(EXCEL_FILE_PATH)
        handler.load()
        workers = handler.get_workers()
        handler.close()

        # Filter out deleted workers
        active_workers = [w for w in workers if w['name'] not in deleted_workers]

        return jsonify({'success': True, 'workers': active_workers})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/workers/all', methods=['GET'])
def get_all_workers_with_proficiency():
    """Get all workers with proficiency matrix"""
    try:
        handler = ExcelHandler(EXCEL_FILE_PATH)
        handler.load()

        machines = handler.get_machines()
        workers = handler.get_workers()
        proficiencies = handler.get_all_proficiencies()

        handler.close()

        return jsonify({
            'success': True,
            'machines': machines,
            'workers': workers,
            'proficiencies': proficiencies
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/workers/update', methods=['POST'])
def update_worker():
    """Update worker name and proficiencies"""
    try:
        data = request.get_json()
        worker_col = data.get('worker_col')
        new_name = data.get('name')
        proficiencies = data.get('proficiencies', {})

        handler = ExcelHandler(EXCEL_FILE_PATH)
        handler.load()

        # Update worker name
        if new_name:
            handler.update_worker_name(worker_col, new_name)

        # Update proficiencies
        if proficiencies:
            prof_data = {}
            for machine_row, prof_value in proficiencies.items():
                if int(machine_row) not in prof_data:
                    prof_data[int(machine_row)] = {}
                prof_data[int(machine_row)][worker_col] = prof_value

            handler.update_proficiencies_bulk(prof_data)

        handler.close()

        return jsonify({'success': True, 'message': 'Worker updated successfully'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/proficiency', methods=['GET'])
def get_proficiency():
    """API endpoint to get proficiency for a machine-worker pair"""
    try:
        machine_row = int(request.args.get('machine_row'))
        worker_col = int(request.args.get('worker_col'))

        handler = ExcelHandler(EXCEL_FILE_PATH)
        handler.load()
        proficiency = handler.get_proficiency(machine_row, worker_col)
        handler.close()

        return jsonify({'success': True, 'proficiency': proficiency})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/schedule', methods=['GET'])
def get_schedule():
    """Get schedule for a specific date"""
    try:
        date_str = request.args.get('date')
        schedule_date = datetime.strptime(date_str, '%Y-%m-%d').date()

        handler = ExcelHandler(EXCEL_FILE_PATH)
        handler.load()
        schedule = handler.get_schedule(schedule_date)
        handler.close()

        return jsonify({'success': True, 'schedule': schedule})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/schedule', methods=['POST'])
def save_schedule():
    """Save schedule with visual formatting"""
    try:
        data = request.get_json()
        date_str = data.get('date')
        schedule_data = data.get('schedule')

        schedule_date = datetime.strptime(date_str, '%Y-%m-%d').date()

        handler = ExcelHandler(EXCEL_FILE_PATH)
        handler.load()
        handler.save_schedule_visual(schedule_date, schedule_data)
        handler.close()

        return jsonify({'success': True, 'message': 'Schedule saved successfully with formatting'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/workers/delete', methods=['POST'])
def delete_worker():
    """Soft delete a worker from future schedules"""
    try:
        data = request.get_json()
        worker_name = data.get('worker_name')

        deleted_workers.add(worker_name)
        save_deleted_workers()

        return jsonify({'success': True, 'message': f'Worker {worker_name} removed from future schedules'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/workers/restore', methods=['POST'])
def restore_worker():
    """Restore a deleted worker"""
    try:
        data = request.get_json()
        worker_name = data.get('worker_name')

        if worker_name in deleted_workers:
            deleted_workers.remove(worker_name)
            save_deleted_workers()

        return jsonify({'success': True, 'message': f'Worker {worker_name} restored'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/workers/deleted', methods=['GET'])
def get_deleted_workers():
    """Get list of deleted workers"""
    return jsonify({'success': True, 'workers': list(deleted_workers)})

if __name__ == '__main__':
    load_deleted_workers()
    app.run(debug=True, host='0.0.0.0', port=5000)
''',

    'templates/index.html': '''<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Machine Schedule Manager</title>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
    <link href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.4.0/css/all.min.css" rel="stylesheet">
    <link href="{{ url_for('static', filename='css/style.css') }}" rel="stylesheet">
</head>
<body>
    <nav class="navbar navbar-dark bg-primary">
        <div class="container-fluid">
            <span class="navbar-brand mb-0 h1">
                <i class="fas fa-calendar-alt"></i> Machine Schedule Manager
            </span>
            <div>
                <a href="/workers" class="btn btn-light btn-sm me-2">
                    <i class="fas fa-user-edit"></i> Edit Workers
                </a>
                <a href="/admin" class="btn btn-light btn-sm">
                    <i class="fas fa-users-cog"></i> Manage Workers
                </a>
            </div>
        </div>
    </nav>

    <div class="container-fluid mt-4">
        <div class="row mb-3">
            <div class="col-md-4">
                <label for="scheduleDate" class="form-label">Select Date:</label>
                <input type="date" id="scheduleDate" class="form-control" value="">
            </div>
            <div class="col-md-4 d-flex align-items-end">
                <button id="loadSchedule" class="btn btn-primary">
                    <i class="fas fa-download"></i> Load Schedule
                </button>
            </div>
        </div>

        <div class="row">
            <div class="col-12">
                <div class="card">
                    <div class="card-header bg-secondary text-white">
                        <h5 class="mb-0">Schedule Assignment</h5>
                    </div>
                    <div class="card-body">
                        <div id="scheduleContainer">
                            <!-- Schedule will be loaded here -->
                        </div>
                    </div>
                </div>
            </div>
        </div>

        <div class="row mt-3">
            <div class="col-12">
                <button id="saveSchedule" class="btn btn-success btn-lg">
                    <i class="fas fa-save"></i> Save Schedule
                </button>
            </div>
        </div>
    </div>

    <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/js/bootstrap.bundle.min.js"></script>
    <script src="{{ url_for('static', filename='js/app.js') }}"></script>
</body>
</html>
''',

    'templates/admin.html': '''<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Worker Management</title>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
    <link href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.4.0/css/all.min.css" rel="stylesheet">
    <link href="{{ url_for('static', filename='css/style.css') }}" rel="stylesheet">
</head>
<body>
    <nav class="navbar navbar-dark bg-primary">
        <div class="container-fluid">
            <span class="navbar-brand mb-0 h1">
                <i class="fas fa-users-cog"></i> Worker Management
            </span>
            <div>
                <a href="/" class="btn btn-light btn-sm">
                    <i class="fas fa-arrow-left"></i> Back to Schedule
                </a>
            </div>
        </div>
    </nav>

    <div class="container mt-4">
        <div class="row">
            <div class="col-md-6">
                <div class="card">
                    <div class="card-header bg-success text-white">
                        <h5 class="mb-0">Active Workers</h5>
                    </div>
                    <div class="card-body">
                        <ul id="activeWorkersList" class="list-group">
                            <!-- Active workers will be loaded here -->
                        </ul>
                    </div>
                </div>
            </div>
            <div class="col-md-6">
                <div class="card">
                    <div class="card-header bg-danger text-white">
                        <h5 class="mb-0">Deleted Workers</h5>
                    </div>
                    <div class="card-body">
                        <ul id="deletedWorkersList" class="list-group">
                            <!-- Deleted workers will be loaded here -->
                        </ul>
                    </div>
                </div>
            </div>
        </div>
    </div>

    <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/js/bootstrap.bundle.min.js"></script>
    <script src="{{ url_for('static', filename='js/admin.js') }}"></script>
</body>
</html>
''',

    'templates/workers.html': '''<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Edit Workers & Proficiency</title>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
    <link href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.4.0/css/all.min.css" rel="stylesheet">
    <style>
        .proficiency-matrix {
            overflow-x: auto;
        }
        .proficiency-matrix table {
            font-size: 0.85rem;
        }
        .proficiency-cell {
            min-width: 120px;
        }
        .machine-name {
            font-weight: bold;
            background-color: #f8f9fa;
        }
        .worker-header {
            background-color: #e9ecef;
            font-weight: bold;
            writing-mode: vertical-rl;
            text-orientation: mixed;
            min-height: 150px;
        }
        .editable-name {
            border: 1px solid #ced4da;
            padding: 2px 5px;
            border-radius: 3px;
            cursor: pointer;
        }
        .editable-name:hover {
            background-color: #f8f9fa;
        }
    </style>
</head>
<body>
    <nav class="navbar navbar-dark bg-primary">
        <div class="container-fluid">
            <span class="navbar-brand mb-0 h1">
                <i class="fas fa-user-edit"></i> Edit Workers & Proficiency
            </span>
            <div>
                <a href="/" class="btn btn-light btn-sm">
                    <i class="fas fa-arrow-left"></i> Back to Schedule
                </a>
                <a href="/admin" class="btn btn-light btn-sm ms-2">
                    <i class="fas fa-users-cog"></i> Manage Workers
                </a>
            </div>
        </div>
    </nav>

    <div class="container-fluid mt-4">
        <div class="row mb-3">
            <div class="col-12">
                <div class="alert alert-info">
                    <i class="fas fa-info-circle"></i>
                    <strong>Instructions:</strong> Click on worker names to edit. Select proficiency level for each machine-worker combination.
                    Proficiency levels: <strong>MR</strong> = Main Role, <strong>C</strong> = Competent, <strong>T</strong> = Trainee, <strong>-</strong> = None
                </div>
            </div>
        </div>

        <div class="row">
            <div class="col-12">
                <div class="card">
                    <div class="card-header bg-secondary text-white">
                        <h5 class="mb-0">Proficiency Matrix</h5>
                    </div>
                    <div class="card-body proficiency-matrix">
                        <table class="table table-bordered table-sm" id="proficiencyTable">
                            <thead>
                                <tr id="workerHeaders">
                                    <th class="machine-name">Machine / Worker</th>
                                    <!-- Worker headers will be added here -->
                                </tr>
                            </thead>
                            <tbody id="proficiencyBody">
                                <!-- Proficiency rows will be added here -->
                            </tbody>
                        </table>
                    </div>
                </div>
            </div>
        </div>

        <div class="row mt-3">
            <div class="col-12">
                <button id="saveChanges" class="btn btn-success btn-lg">
                    <i class="fas fa-save"></i> Save All Changes
                </button>
            </div>
        </div>
    </div>

    <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/js/bootstrap.bundle.min.js"></script>
    <script src="{{ url_for('static', filename='js/workers.js') }}"></script>
</body>
</html>
''',

    'static/css/style.css': '''body {
    background-color: #f5f5f5;
}

.machine-card {
    transition: all 0.3s ease;
}

.machine-card:hover {
    box-shadow: 0 4px 8px rgba(0,0,0,0.1);
}

.machine-details {
    background-color: #f8f9fa;
    padding: 15px;
}

.worker-item {
    padding: 8px;
    border-left: 3px solid #007bff;
    margin 'static/css/style.css': ''body {
    background-color: #f5f5f5;
}

.machine-card {
    transition: all 0.3s ease;
}

.machine-card:hover {
    box-shadow: 0 4px 8px rgba(0,0,0,0.1);
}

.machine-details {
    background-color: #f8f9fa;
    padding: 15px;
}

.worker-item {
    padding: 8px;
    border-left: 3px solid #007bff;
    margin-bottom: 8px;
    background-color: white;
}

.role-selection {
    margin-top: 8px;
}

.btn-group-sm .btn {
    font-size: 0.75rem;
    padding: 0.25rem 0.5rem;
}

.alert {
    border-radius: 8px;
}

.card {
    border-radius: 8px;
    box-shadow: 0 2px 4px rgba(0,0,0,0.1);
}

.card-header {
    border-radius: 8px 8px 0 0 !important;
}

@media print {
    .navbar, .btn, .alert {
        display: none;
    }
}
''',

    'static/js/app.js': '''let machines = [];
let workers = [];
let currentSchedule = {};

// Set today's date as default
document.addEventListener('DOMContentLoaded', function() {
    const today = new Date().toISOString().split('T')[0];
    document.getElementById('scheduleDate').value = today;
    
    loadMachinesAndWorkers();
    
    document.getElementById('loadSchedule').addEventListener('click', loadSchedule);
    document.getElementById('saveSchedule').addEventListener('click', saveSchedule);
});

async function loadMachinesAndWorkers() {
    try {
        // Load machines
        const machinesResponse = await fetch('/api/machines');
        const machinesData = await machinesResponse.json();
        
        if (machinesData.success) {
            machines = machinesData.machines;
        }
        
        // Load workers
        const workersResponse = await fetch('/api/workers');
        const workersData = await workersResponse.json();
        
        if (workersData.success) {
            workers = workersData.workers;
        }
        
        renderScheduleInterface();
    } catch (error) {
        console.error('Error loading data:', error);
        alert('Error loading machines and workers');
    }
}

function renderScheduleInterface() {
    const container = document.getElementById('scheduleContainer');
    container.innerHTML = '';
    
    machines.forEach(machine => {
        const machineCard = document.createElement('div');
        machineCard.className = 'machine-card mb-3';
        machineCard.innerHTML = `
            <div class="card">
                <div class="card-header">
                    <div class="form-check">
                        <input class="form-check-input machine-checkbox" type="checkbox" 
                               id="machine-${machine.row}" data-machine-row="${machine.row}" 
                               data-machine-name="${machine.name}">
                        <label class="form-check-label" for="machine-${machine.row}">
                            <strong>${machine.name}</strong>
                        </label>
                    </div>
                </div>
                <div class="card-body machine-details" id="details-${machine.row}" style="display:none;">
                    <div class="row">
                        <div class="col-md-6">
                            <h6>Assign Workers:</h6>
                            <div id="workers-${machine.row}">
                                ${renderWorkersList(machine)}
                            </div>
                        </div>
                        <div class="col-md-6">
                            <h6>Time:</h6>
                            <div class="mb-2">
                                <label class="form-label">Start Time:</label>
                                <input type="time" class="form-control time-start" 
                                       data-machine-row="${machine.row}">
                            </div>
                            <div class="mb-2">
                                <label class="form-label">Finish Time:</label>
                                <input type="time" class="form-control time-finish" 
                                       data-machine-row="${machine.row}">
                            </div>
                        </div>
                    </div>
                </div>
            </div>
        `;
        
        container.appendChild(machineCard);
        
        // Add event listener for checkbox
        const checkbox = machineCard.querySelector(`#machine-${machine.row}`);
        checkbox.addEventListener('change', function() {
            const details = document.getElementById(`details-${machine.row}`);
            details.style.display = this.checked ? 'block' : 'none';
        });
    });
}

function renderWorkersList(machine) {
    let html = '<div class="worker-list">';
    
    workers.forEach(worker => {
        html += `
            <div class="worker-item mb-2">
                <div class="form-check">
                    <input class="form-check-input worker-checkbox" type="checkbox" 
                           id="worker-${machine.row}-${worker.col}" 
                           data-machine-row="${machine.row}"
                           data-worker-name="${worker.name}"
                           data-worker-col="${worker.col}">
                    <label class="form-check-label" for="worker-${machine.row}-${worker.col}">
                        ${worker.name}
                    </label>
                </div>
                <div class="role-selection ms-4" id="role-${machine.row}-${worker.col}" style="display:none;">
                    <div class="btn-group btn-group-sm" role="group">
                        <input type="radio" class="btn-check" name="role-${machine.row}-${worker.col}" 
                               id="main-${machine.row}-${worker.col}" value="Main Role">
                        <label class="btn btn-outline-primary" for="main-${machine.row}-${worker.col}">Main Role</label>
                        
                        <input type="radio" class="btn-check" name="role-${machine.row}-${worker.col}" 
                               id="competent-${machine.row}-${worker.col}" value="Competent">
                        <label class="btn btn-outline-success" for="competent-${machine.row}-${worker.col}">Competent</label>
                        
                        <input type="radio" class="btn-check" name="role-${machine.row}-${worker.col}" 
                               id="trainee-${machine.row}-${worker.col}" value="Trainee">
                        <label class="btn btn-outline-warning" for="trainee-${machine.row}-${worker.col}">Trainee</label>
                    </div>
                </div>
            </div>
        `;
    });
    
    html += '</div>';
    
    // Add event listeners after rendering
    setTimeout(() => {
        workers.forEach(worker => {
            const checkbox = document.getElementById(`worker-${machine.row}-${worker.col}`);
            if (checkbox) {
                checkbox.addEventListener('change', function() {
                    const roleDiv = document.getElementById(`role-${machine.row}-${worker.col}`);
                    roleDiv.style.display = this.checked ? 'block' : 'none';
                });
            }
        });
    }, 100);
    
    return html;
}

async function loadSchedule() {
    const dateInput = document.getElementById('scheduleDate').value;
    if (!dateInput) {
        alert('Please select a date');
        return;
    }
    
    try {
        const response = await fetch(`/api/schedule?date=${dateInput}`);
        const data = await response.json();
        
        if (data.success) {
            applyScheduleToInterface(data.schedule);
        }
    } catch (error) {
        console.error('Error loading schedule:', error);
        alert('Error loading schedule');
    }
}

function applyScheduleToInterface(schedule) {
    // Reset all selections
    document.querySelectorAll('.machine-checkbox').forEach(cb => cb.checked = false);
    document.querySelectorAll('.worker-checkbox').forEach(cb => cb.checked = false);
    document.querySelectorAll('.machine-details').forEach(div => div.style.display = 'none');
    
    // Group schedule by machine
    const scheduleByMachine = {};
    schedule.forEach(entry => {
        if (!scheduleByMachine[entry.machine]) {
            scheduleByMachine[entry.machine] = {
                workers: [],
                time_start: entry.time_start,
                time_finish: entry.time_finish
            };
        }
        scheduleByMachine[entry.machine].workers.push({
            name: entry.worker,
            role: entry.role
        });
    });
    
    // Apply to interface
    Object.keys(scheduleByMachine).forEach(machineName => {
        const machine = machines.find(m => m.name === machineName);
        if (machine) {
            // Check machine
            const machineCheckbox = document.querySelector(`[data-machine-name="${machineName}"]`);
            if (machineCheckbox) {
                machineCheckbox.checked = true;
                machineCheckbox.dispatchEvent(new Event('change'));
                
                // Set times
                document.querySelector(`.time-start[data-machine-row="${machine.row}"]`).value = 
                    scheduleByMachine[machineName].time_start || '';
                document.querySelector(`.time-finish[data-machine-row="${machine.row}"]`).value = 
                    scheduleByMachine[machineName].time_finish || '';
                
                // Set workers and roles
                scheduleByMachine[machineName].workers.forEach(workerEntry => {
                    const worker = workers.find(w => w.name === workerEntry.name);
                    if (worker) {
                        const workerCheckbox = document.getElementById(`worker-${machine.row}-${worker.col}`);
                        if (workerCheckbox) {
                            workerCheckbox.checked = true;
                            workerCheckbox.dispatchEvent(new Event('change'));
                            
                            // Set role
                            const roleRadio = document.querySelector(
                                `input[name="role-${machine.row}-${worker.col}"][value="${workerEntry.role}"]`
                            );
                            if (roleRadio) {
                                roleRadio.checked = true;
                            }
                        }
                    }
                });
            }
        }
    });
}

async function saveSchedule() {
    const dateInput = document.getElementById('scheduleDate').value;
    if (!dateInput) {
        alert('Please select a date');
        return;
    }
    
    const scheduleData = [];
    
    // Iterate through checked machines
    document.querySelectorAll('.machine-checkbox:checked').forEach(machineCheckbox => {
        const machineRow = machineCheckbox.dataset.machineRow;
        const machineName = machineCheckbox.dataset.machineName;
        
        const timeStart = document.querySelector(`.time-start[data-machine-row="${machineRow}"]`).value;
        const timeFinish = document.querySelector(`.time-finish[data-machine-row="${machineRow}"]`).value;
        
        // Get selected workers for this machine
        document.querySelectorAll(`.worker-checkbox[data-machine-row="${machineRow}"]:checked`).forEach(workerCheckbox => {
            const workerName = workerCheckbox.dataset.workerName;
            const workerCol = workerCheckbox.dataset.workerCol;
            
            // Get selected role
            const roleRadio = document.querySelector(
                `input[name="role-${machineRow}-${workerCol}"]:checked`
            );
            const role = roleRadio ? roleRadio.value : '';
            
            scheduleData.push({
                machine: machineName,
                worker: workerName,
                role: role,
                time_start: timeStart,
                time_finish: timeFinish
            });
        });
    });
    
    try {
        const response = await fetch('/api/schedule', {
            method: 'POST',
            headers: {
                'Content-Type': 'application/json'
            },
            body: JSON.stringify({
                date: dateInput,
                schedule: scheduleData
            })
        });
        
        const data = await response.json();
        
        if (data.success) {
            alert('Schedule saved successfully!');
        } else {
            alert('Error saving schedule: ' + data.error);
        }
    } catch (error) {
        console.error('Error saving schedule:', error);
        alert('Error saving schedule');
    }
}
''',

    'static/js/admin.js': '''let allWorkers = [];
let deletedWorkers = [];

document.addEventListener('DOMContentLoaded', function() {
    loadWorkers();
});

async function loadWorkers() {
    try {
        // Load all workers (from Excel)
        const allResponse = await fetch('/api/workers/all');
        const allData = await allResponse.json();
        
        if (allData.success) {
            allWorkers = allData.workers;
        }
        
        // Load deleted workers list
        const deletedResponse = await fetch('/api/workers/deleted');
        const deletedData = await deletedResponse.json();
        
        if (deletedData.success) {
            deletedWorkers = deletedData.workers;
        }
        
        renderWorkerLists();
    } catch (error) {
        console.error('Error loading workers:', error);
        alert('Error loading workers');
    }
}

function renderWorkerLists() {
    const activeList = document.getElementById('activeWorkersList');
    const deletedList = document.getElementById('deletedWorkersList');
    
    activeList.innerHTML = '';
    deletedList.innerHTML = '';
    
    // Render active workers
    allWorkers.forEach(worker => {
        if (!deletedWorkers.includes(worker.name)) {
            const li = document.createElement('li');
            li.className = 'list-group-item d-flex justify-content-between align-items-center';
            li.innerHTML = `
                <span>${worker.name}</span>
                <button class="btn btn-danger btn-sm" onclick="deleteWorker('${worker.name}')">
                    <i class="fas fa-trash"></i> Delete
                </button>
            `;
            activeList.appendChild(li);
        }
    });
    
    // Render deleted workers
    deletedWorkers.forEach(workerName => {
        const li = document.createElement('li');
        li.className = 'list-group-item d-flex justify-content-between align-items-center';
        li.innerHTML = `
            <span>${workerName}</span>
            <button class="btn btn-success btn-sm" onclick="restoreWorker('${workerName}')">
                <i class="fas fa-undo"></i> Restore
            </button>
        `;
        deletedList.appendChild(li);
    });
    
    if (activeList.children.length === 0) {
        activeList.innerHTML = '<li class="list-group-item text-muted">No active workers</li>';
    }
    
    if (deletedList.children.length === 0) {
        deletedList.innerHTML = '<li class="list-group-item text-muted">No deleted workers</li>';
    }
}

async function deleteWorker(workerName) {
    if (!confirm(`Are you sure you want to remove ${workerName} from future schedules?`)) {
        return;
    }
    
    try {
        const response = await fetch('/api/workers/delete', {
            method: 'POST',
            headers: {
                'Content-Type': 'application/json'
            },
            body: JSON.stringify({
                worker_name: workerName
            })
        });
        
        const data = await response.json();
        
        if (data.success) {
            deletedWorkers.push(workerName);
            renderWorkerLists();
        } else {
            alert('Error deleting worker: ' + data.error);
        }
    } catch (error) {
        console.error('Error deleting worker:', error);
        alert('Error deleting worker');
    }
}

async function restoreWorker(workerName) {
    try {
        const response = await fetch('/api/workers/restore', {
            method: 'POST',
            headers: {
                'Content-Type': 'application/json'
            },
            body: JSON.stringify({
                worker_name: workerName
            })
        });
        
        const data = await response.json();
        
        if (data.success) {
            deletedWorkers = deletedWorkers.filter(w => w !== workerName);
            renderWorkerLists();
        } else {
            alert('Error restoring worker: ' + data.error);
        }
    } catch (error) {
        console.error('Error restoring worker:', error);
        alert('Error restoring worker');
    }
}
''',

    'static/js/workers.js': '''let machines = [];
let workers = [];
let proficiencies = {};
let changes = {};

document.addEventListener('DOMContentLoaded', function() {
    loadWorkersAndProficiency();
    document.getElementById('saveChanges').addEventListener('click', saveAllChanges);
});

async function loadWorkersAndProficiency() {
    try {
        const response = await fetch('/api/workers/all');
        const data = await response.json();
        
        if (data.success) {
            machines = data.machines;
            workers = data.workers;
            proficiencies = data.proficiencies;
            
            renderProficiencyMatrix();
        } else {
            alert('Error loading data: ' + data.error);
        }
    } catch (error) {
        console.error('Error:', error);
        alert('Error loading workers and proficiency data');
    }
}

function renderProficiencyMatrix() {
    const headerRow = document.getElementById('workerHeaders');
    const tbody = document.getElementById('proficiencyBody');
    
    // Clear existing content
    headerRow.innerHTML = '<th class="machine-name">Machine / Worker</th>';
    tbody.innerHTML = '';
    
    // Add worker headers
    workers.forEach(worker => {
        const th = document.createElement('th');
        th.className = 'worker-header text-center';
        th.innerHTML = `
            <div class="editable-name" 
                 data-worker-col="${worker.col}" 
                 onclick="editWorkerName(this, '${worker.name}', ${worker.col})">
                ${worker.name}
            </div>
        `;
        headerRow.appendChild(th);
    });
    
    // Add machine rows
    machines.forEach(machine => {
        const row = document.createElement('tr');
        
        // Machine name cell
        const machineCell = document.createElement('td');
        machineCell.className = 'machine-name';
        machineCell.textContent = machine.name;
        row.appendChild(machineCell);
        
        // Proficiency cells for each worker
        workers.forEach(worker => {
            const cell = document.createElement('td');
            cell.className = 'proficiency-cell text-center';
            
            const currentProf = proficiencies[machine.row] && proficiencies[machine.row][worker.col] 
                ? proficiencies[machine.row][worker.col] 
                : '';
            
            cell.innerHTML = `
                <select class="form-select form-select-sm proficiency-select" 
                        data-machine-row="${machine.row}" 
                        data-worker-col="${worker.col}"
                        onchange="trackChange(${machine.row}, ${worker.col}, this.value)">
                    <option value="" ${currentProf === '' ? 'selected' : ''}>-</option>
                    <option value="Main Role" ${currentProf === 'Main Role' ? 'selected' : ''}>Main Role (MR)</option>
                    <option value="Competent" ${currentProf === 'Competent' ? 'selected' : ''}>Competent (C)</option>
                    <option value="Trainee" ${currentProf === 'Trainee' ? 'selected' : ''}>Trainee (T)</option>
                </select>
            `;
            
            row.appendChild(cell);
        });
        
        tbody.appendChild(row);
    });
}

function editWorkerName(element, currentName, workerCol) {
    const newName = prompt('Enter new name for worker:', currentName);
    
    if (newName && newName !== currentName) {
        element.textContent = newName;
        
        // Track name change
        if (!changes['names']) {
            changes['names'] = {};
        }
        changes['names'][workerCol] = newName;
        
        // Update workers array
        const worker = workers.find(w => w.col === workerCol);
        if (worker) {
            worker.name = newName;
        }
    }
}

function trackChange(machineRow, workerCol, proficiency) {
    if (!changes['proficiencies']) {
        changes['proficiencies'] = {};
    }
    
    if (!changes['proficiencies'][workerCol]) {
        changes['proficiencies'][workerCol] = {};
    }
    
    changes['proficiencies'][workerCol][machineRow] = proficiency;
}

async function saveAllChanges() {
    if (Object.keys(changes).length === 0) {
        alert('No changes to save');
        return;
    }
    
    const saveButton = document.getElementById('saveChanges');
    saveButton.disabled = true;
    saveButton.innerHTML = '<i class="fas fa-spinner fa-spin"></i> Saving...';
    
    try {
        // Save each worker's changes
        const promises = [];
        
        if (changes['names']) {
            for (const [workerCol, newName] of Object.entries(changes['names'])) {
                const profData = changes['proficiencies'] && changes['proficiencies'][workerCol] 
                    ? changes['proficiencies'][workerCol] 
                    : {};
                
                promises.push(
                    fetch('/api/workers/update', {
                        method: 'POST',
                        headers: { 'Content-Type': 'application/json' },
                        body: JSON.stringify({
                            worker_col: parseInt(workerCol),
                            name: newName,
                            proficiencies: profData
                        })
                    })
                );
            }
        }
        
        // Save proficiency changes for workers without name changes
        if (changes['proficiencies']) {
            for (const [workerCol, profData] of Object.entries(changes['proficiencies'])) {
                if (!changes['names'] || !changes['names'][workerCol]) {
                    promises.push(
                        fetch('/api/workers/update', {
                            method: 'POST',
                            headers: { 'Content-Type': 'application/json' },
                            body: JSON.stringify({
                                worker_col: parseInt(workerCol),
                                proficiencies: profData
                            })
                        })
                    );
                }
            }
        }
        
        await Promise.all(promises);
        
        alert('All changes saved successfully!');
        changes = {};
        
        // Reload data
        await loadWorkersAndProficiency();
        
    } catch (error) {
        console.error('Error saving changes:', error);
        alert('Error saving changes');
    } finally {
        saveButton.disabled = false;
        saveButton.innerHTML = '<i class="fas fa-save"></i> Save All Changes';
    }
}
'''
}

# Write all files
print("\nCreating files...")
for filename, content in files.items():
    with open(filename, 'w', encoding='utf-8') as f:
        f.write(content)
    print(f'✓ Created: {filename}')

print('\n' + '='*60)
print('✓ PROJECT SETUP COMPLETE!')
print('='*60)
print('\nNext steps:')
print('1. Install dependencies:')
print('   pip install -r requirements.txt')
print('\n2. Run the application:')
print('   python app.py')
print('\n3. Open your browser:')
print('   http://localhost:5000')
print('\n4. Make sure your Excel file exists at:')
print('   C:\\Users\\mtheu\\Downloads\\test.xlsx')
print('\nPages available:')
print('  - Main Schedule: http://localhost:5000/')
print('  - Edit Workers: http://localhost:5000/workers')
print('  - Manage Workers: http://localhost:5000/admin')
print('='*60)