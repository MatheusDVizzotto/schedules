# create_valid_template.py - Create and upload valid Excel template
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
from google_drive_handler import GoogleDriveHandler
import io


def create_valid_excel():
    """Create a valid Excel workbook"""
    print("Creating Excel workbook...")

    wb = Workbook()
    ws = wb.active
    ws.title = "Main"

    # Add machines in specified ranges
    print("Adding machines...")

    # A5:A8
    machines_1 = ['CNC Machine 1', 'CNC Machine 2', 'Lathe 1', 'Lathe 2']
    for i, machine in enumerate(machines_1, start=5):
        ws[f'A{i}'] = machine
        print(f"  Added: {machine} at A{i}")

    # A10:A22
    machines_2 = ['Mill 1', 'Mill 2', 'Drill Press 1', 'Drill Press 2',
                  'Grinder 1', 'Grinder 2', 'Press 1', 'Press 2',
                  'Welder 1', 'Welder 2', 'Assembly 1', 'Assembly 2', 'Quality Control']
    for i, machine in enumerate(machines_2, start=10):
        ws[f'A{i}'] = machine
        print(f"  Added: {machine} at A{i}")

    # A24:A36
    machines_3 = ['Packaging 1', 'Packaging 2', 'Shipping 1', 'Shipping 2',
                  'Maintenance 1', 'Maintenance 2', 'Storage 1', 'Storage 2',
                  'Office 1', 'Office 2', 'Break Room', 'Training Room', 'Conference Room']
    for i, machine in enumerate(machines_3, start=24):
        ws[f'A{i}'] = machine
        print(f"  Added: {machine} at A{i}")

    # Add workers in columns AO-AX (41-50)
    print("\nAdding workers...")
    workers = ['John Smith', 'Jane Doe', 'Bob Johnson', 'Alice Williams',
               'Charlie Brown', 'Diana Prince', 'Eve Martinez', 'Frank Garcia',
               'Grace Lee', 'Henry Wilson']

    for col_num, worker in enumerate(workers, start=41):
        col_letter = get_column_letter(col_num)
        ws[f'{col_letter}1'] = worker
        print(f"  Added: {worker} at {col_letter}1")

    # Add some proficiency data
    print("\nAdding proficiency data...")
    # Just add a few examples
    ws['AO5'] = 'Main Role'  # John Smith - CNC Machine 1
    ws['AP5'] = 'Competent'  # Jane Doe - CNC Machine 1
    ws['AO6'] = 'Competent'  # John Smith - CNC Machine 2
    ws['AP6'] = 'Main Role'  # Jane Doe - CNC Machine 2

    # Add formatting
    print("Adding formatting...")

    # Header fill
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col in range(41, 51):
        col_letter = get_column_letter(col)
        cell = ws[f'{col_letter}1']
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Machine formatting
    machine_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    machine_font = Font(bold=True)

    for row in [5, 6, 7, 8] + list(range(10, 23)) + list(range(24, 37)):
        if ws[f'A{row}'].value:
            ws[f'A{row}'].fill = machine_fill
            ws[f'A{row}'].font = machine_font

    # Set column widths
    ws.column_dimensions['A'].width = 20
    for col in range(41, 51):
        ws.column_dimensions[get_column_letter(col)].width = 15

    print("\n✓ Excel workbook created successfully")
    return wb


def test_workbook_validity(wb):
    """Test if workbook can be saved and loaded"""
    print("\nTesting workbook validity...")
    try:
        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Try to load it back
        from openpyxl import load_workbook
        test_wb = load_workbook(buffer)

        print("✓ Workbook is valid and can be loaded")
        return True
    except Exception as e:
        print(f"✗ Workbook validation failed: {e}")
        return False


def upload_to_gdrive(wb, filename='schedule.xlsx'):
    """Upload workbook to Google Drive"""
    print(f"\n{'=' * 60}")
    print("Uploading to Google Drive")
    print('=' * 60)

    try:
        # Connect to Google Drive
        print("Connecting to Google Drive...")
        gdrive = GoogleDriveHandler()
        print("✓ Connected")

        # Check if file exists
        print(f"\nChecking if '{filename}' exists...")
        existing_id = gdrive.get_file_id_by_name(filename)

        if existing_id:
            print(f"⚠ File already exists (ID: {existing_id})")
            choice = input("  Overwrite? (yes/no): ")
            if choice.lower() != 'yes':
                print("Upload cancelled")
                return None

        # Save to buffer
        print("\nPreparing file...")
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Upload
        if existing_id:
            print("Updating existing file...")
            gdrive.upload_file(existing_id, buffer)
            file_id = existing_id
        else:
            print("Creating new file...")
            file_id = gdrive.create_file(filename, buffer)

        if file_id:
            print(f"\n{'=' * 60}")
            print("✓ SUCCESS!")
            print('=' * 60)
            print(f"File ID: {file_id}")
            print(f"Filename: {filename}")

            # Get link
            metadata = gdrive.get_file_metadata(file_id)
            if metadata and 'webViewLink' in metadata:
                print(f"View: {metadata['webViewLink']}")

            # Update config
            print("\nUpdating config.py...")
            update_config(file_id, filename)

            return file_id
        else:
            print("\n✗ Upload failed")
            return None

    except Exception as e:
        print(f"\n✗ Error: {e}")
        import traceback
        traceback.print_exc()
        return None


def update_config(file_id, filename):
    """Update config.py with file details"""
    try:
        with open('config.py', 'r') as f:
            content = f.read()

        import re

        # Update file ID
        content = re.sub(
            r"GOOGLE_DRIVE_FILE_ID = ['\"].*?['\"]",
            f"GOOGLE_DRIVE_FILE_ID = '{file_id}'",
            content
        )

        # Update filename
        content = re.sub(
            r"GOOGLE_DRIVE_FILENAME = ['\"].*?['\"]",
            f"GOOGLE_DRIVE_FILENAME = '{filename}'",
            content
        )

        # Set to use file ID
        content = re.sub(
            r"USE_FILE_ID = (True|False)",
            "USE_FILE_ID = True",
            content
        )

        with open('config.py', 'w') as f:
            f.write(content)

        print("✓ config.py updated")

    except Exception as e:
        print(f"⚠ Could not update config.py: {e}")
        print(f"\nPlease manually update config.py:")
        print(f"  GOOGLE_DRIVE_FILE_ID = '{file_id}'")
        print(f"  GOOGLE_DRIVE_FILENAME = '{filename}'")
        print(f"  USE_FILE_ID = True")


if __name__ == '__main__':
    print("=" * 60)
    print("Valid Excel Template Creator for Google Drive")
    print("=" * 60)
    print()

    # Get filename
    filename = input("Enter filename (default: schedule.xlsx): ").strip()
    if not filename:
        filename = 'schedule.xlsx'
    if not filename.endswith('.xlsx'):
        filename += '.xlsx'

    print()

    # Create workbook
    wb = create_valid_excel()

    # Test validity
    if not test_workbook_validity(wb):
        print("\n✗ Workbook creation failed")
        exit(1)

    # Upload
    file_id = upload_to_gdrive(wb, filename)

    if file_id:
        print("\n" + "=" * 60)
        print("READY TO USE!")
        print("=" * 60)
        print("\nYou can now run:")
        print("  python app.py")
        print("\nThen open:")
        print("  http://localhost:5000")
        print("=" * 60)
    else:
        print("\n✗ Setup incomplete")